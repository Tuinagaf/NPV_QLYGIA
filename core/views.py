import time
from django.utils import timezone
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models
from .models import DoiTac, TuyenXe, GiaCoSo, BaoGiaThongTinXe, LichSuBaoGia, DeXuatGiaCoSo
from .forms import DoiTacForm


def clean_huyen(name):
    if not name: return name
    for p in ["Quận ", "Huyện ", "Thị xã ", "Thành phố "]:
        if name.startswith(p):
            tmp = name[len(p):].strip()
            if p == "Quận " and tmp.isdigit():
                return name
            return tmp
    return name

def parse_vn_number(val):
    if not val:
        return 0.0
    val_str = str(val).strip().replace(' ', '')
    if not val_str:
        return 0.0
    val_str = val_str.replace('.', '').replace(',', '.')
    try:
        return float(val_str)
    except ValueError:
        return 0.0

@login_required
def dashboard(request):
    if request.user.is_superuser:
        from django.db.models import Prefetch
        partners = DoiTac.objects.prefetch_related('bao_gia__tuyen').all()
        quotes = BaoGiaThongTinXe.objects.all()
    else:
        partners = DoiTac.objects.filter(nguoi_quan_ly=request.user.username)
        quotes = BaoGiaThongTinXe.objects.filter(doi_tac__in=partners)
        
    context = {
        'total_partners': partners.count(),
        'total_routes': TuyenXe.objects.count(),
        'total_prices': GiaCoSo.objects.count(),
        'total_quotations': quotes.count(),
    }
    return render(request, 'core/dashboard.html', context)

@login_required
def partner_list(request):
    search_query = request.GET.get('search', '').strip()
    
    if request.user.is_superuser:
        from django.db.models import Prefetch
        partners = DoiTac.objects.prefetch_related('bao_gia__tuyen').all()
    else:
        partners = DoiTac.objects.filter(nguoi_quan_ly=request.user.username)
    
    # Tìm kiếm theo tên nhà xe, mã nhà xe, số điện thoại, mã số thuế
    if search_query:
        partners = partners.filter(
            models.Q(ten_nha_xe__icontains=search_query) |
            models.Q(ma_nha_xe__icontains=search_query) |
            models.Q(so_dien_thoai__icontains=search_query) |
            models.Q(ma_so_thue__icontains=search_query) |
            models.Q(nguoi_quan_ly__icontains=search_query)
        )
    
    return render(request, 'core/partner_list.html', {
        'partners': partners,
        'search_query': search_query
    })

def process_routes(request, partner):
    tinh_nhans = request.POST.getlist('tinh_nhan[]')
    huyen_nhans = request.POST.getlist('huyen_nhan[]')
    tinh_giaos = request.POST.getlist('tinh_giao[]')
    huyen_giaos = request.POST.getlist('huyen_giao[]')
    tai_trongs = request.POST.getlist('tai_trong_tan[]')
    so_khois = request.POST.getlist('so_khoi[]')
    kich_thuocs = request.POST.getlist('kich_thuoc[]')
    loai_thungs = request.POST.getlist('loai_thung[]')
    muc_gias = request.POST.getlist('muc_gia_chap_nhan[]')
    ngay_bat_daus = request.POST.getlist('ngay_bat_dau[]')
    ghep_hangs = request.POST.getlist('co_ghep_hang_khong[]')
    qua_tais = request.POST.getlist('co_chiu_qua_tai_khong[]')
    nhieu_diems = request.POST.getlist('co_di_nhieu_diem_khong[]')
    chieu_dis = request.POST.getlist('di_1_hay_2_chieu[]')
    bg_ids = request.POST.getlist('bg_id[]')

    valid_ids = [int(i) for i in bg_ids if i.isdigit()]
    BaoGiaThongTinXe.objects.filter(doi_tac=partner).exclude(id__in=valid_ids).update(is_deleted=True)

    for i in range(len(tinh_nhans)):
        if not tinh_nhans[i] or not tinh_giaos[i]: continue
        
        tinh_n = tinh_nhans[i].replace("Tỉnh ", "").replace("Thành phố ", "").strip()
        huyen_n = huyen_nhans[i] if i < len(huyen_nhans) else ''
        huyen_n = clean_huyen(huyen_n)
        
        tinh_g = tinh_giaos[i].replace("Tỉnh ", "").replace("Thành phố ", "").strip()
        huyen_g = huyen_giaos[i] if i < len(huyen_giaos) else ''
        huyen_g = clean_huyen(huyen_g)
        
        try:
            t_str = tai_trongs[i].lower().replace('tấn', '').replace('t', '').strip() if i < len(tai_trongs) and tai_trongs[i] else '0'
            tai_trong = float(t_str)
        except ValueError:
            tai_trong = 0.0
        so_khoi = so_khois[i] if i < len(so_khois) and so_khois[i] else None
        kich_thuoc = kich_thuocs[i] if i < len(kich_thuocs) else ''
        loai_thung = loai_thungs[i] if i < len(loai_thungs) else ''
        
        muc_gia = parse_vn_number(muc_gias[i]) if i < len(muc_gias) and muc_gias[i] else 0.0
            
        ngay_bat_dau = ngay_bat_daus[i] if i < len(ngay_bat_daus) and ngay_bat_daus[i] else timezone.now().date()
        ghep_hang = True if (i < len(ghep_hangs) and ghep_hangs[i] == '1') else False
        qua_tai = True if (i < len(qua_tais) and qua_tais[i] == '1') else False
        nhieu_diem = True if (i < len(nhieu_diems) and nhieu_diems[i] == '1') else False
        chieu_di = int(chieu_dis[i]) if i < len(chieu_dis) and chieu_dis[i].isdigit() else 1
        
        tuyen_queryset = TuyenXe.objects.filter(tinh_nhan=tinh_n, huyen_nhan=huyen_n, tinh_giao=tinh_g, huyen_giao=huyen_g)
        if tuyen_queryset.exists():
            tuyen = tuyen_queryset.first()
        else:
            code_t = f"T-{tinh_n[:2].upper()}-{tinh_g[:2].upper()}-{int(time.time()*1000) % 100000}"
            tuyen = TuyenXe.objects.create(
                ma_tuyen=code_t,
                tinh_nhan=tinh_n,
                huyen_nhan=huyen_n,
                tinh_giao=tinh_g,
                huyen_giao=huyen_g
            )
        
        BaoGiaThongTinXe.objects.create(
            doi_tac=partner,
            tuyen=tuyen,
            tai_trong_tan=tai_trong,
            so_khoi=so_khoi,
            kich_thuoc=kich_thuoc,
            loai_thung=loai_thung,
            ngay_bat_dau=ngay_bat_dau,
            muc_gia_chap_nhan=muc_gia,
            co_ghep_hang_khong=ghep_hang,
            co_chiu_qua_tai_khong=qua_tai,
            co_di_nhieu_diem_khong=nhieu_diem,
            di_1_hay_2_chieu=chieu_di
        )

@login_required
def partner_create(request):
    if request.method == 'POST':
        form = DoiTacForm(request.POST, user=request.user)
        if form.is_valid():
            partner = form.save(commit=False)
            if request.user.is_authenticated and not request.user.is_superuser:
                partner.nguoi_quan_ly = request.user.username
            partner.save()
            process_routes(request, partner)
            messages.success(request, f'Đã tạo đối tác "{partner.ten_nha_xe}" thành công!')
            return redirect('partner_detail', pk=partner.pk)
        else:
            messages.error(request, 'Có lỗi trong form. Vui lòng kiểm tra lại.')
            
            # Giữ lại dữ liệu tuyến mới đã nhập
            tinh_nhans = request.POST.getlist('tinh_nhan[]')
            huyen_nhans = request.POST.getlist('huyen_nhan[]')
            tinh_giaos = request.POST.getlist('tinh_giao[]')
            huyen_giaos = request.POST.getlist('huyen_giao[]')
            tai_trongs = request.POST.getlist('tai_trong_tan[]')
            so_khois = request.POST.getlist('so_khoi[]')
            kich_thuocs = request.POST.getlist('kich_thuoc[]')
            loai_thungs = request.POST.getlist('loai_thung[]')
            muc_gias = request.POST.getlist('muc_gia_chap_nhan[]')
            ghep_hangs = request.POST.getlist('co_ghep_hang_khong[]')
            qua_tais = request.POST.getlist('co_chiu_qua_tai_khong[]')
            nhieu_diems = request.POST.getlist('co_di_nhieu_diem_khong[]')
            chieu_dis = request.POST.getlist('di_1_hay_2_chieu[]')
            
            new_routes_data = []
            for i in range(len(tinh_nhans)):
                new_routes_data.append({
                    'tinh_nhan': tinh_nhans[i],
                    'huyen_nhan': huyen_nhans[i] if i < len(huyen_nhans) else '',
                    'tinh_giao': tinh_giaos[i] if i < len(tinh_giaos) else '',
                    'huyen_giao': huyen_giaos[i] if i < len(huyen_giaos) else '',
                    'tai_trong_tan': tai_trongs[i] if i < len(tai_trongs) else '',
                    'so_khoi': so_khois[i] if i < len(so_khois) else '',
                    'kich_thuoc': kich_thuocs[i] if i < len(kich_thuocs) else '',
                    'loai_thung': loai_thungs[i] if i < len(loai_thungs) else '',
                    'muc_gia_chap_nhan': muc_gias[i] if i < len(muc_gias) else '',
                    'co_ghep_hang_khong': ghep_hangs[i] if i < len(ghep_hangs) else '0',
                    'co_chiu_qua_tai_khong': qua_tais[i] if i < len(qua_tais) else '0',
                    'co_di_nhieu_diem_khong': nhieu_diems[i] if i < len(nhieu_diems) else '0',
                    'di_1_hay_2_chieu': chieu_dis[i] if i < len(chieu_dis) else '1'
                })
            
            import json
            return render(request, 'core/partner_form.html', {'form': form, 'title': 'Thêm Đối Tác Mới', 'routes': [], 'new_routes_data_json': json.dumps(new_routes_data)})
            
    else:
        # Người quản lý sẽ tự động được gán là request.user khi form.save()
        initial_data = {}
        if request.user.is_authenticated:
            initial_data['nguoi_quan_ly'] = request.user.username
        form = DoiTacForm(initial=initial_data)
    return render(request, 'core/partner_form.html', {'form': form, 'title': 'Thêm Đối Tác Mới', 'routes': []})

@login_required
def partner_detail(request, pk):
    partner = get_object_or_404(DoiTac, pk=pk)
    
    routes = partner.bao_gia.all()
    
    unique_routes = set()
    unique_types = set()
    
    for r in routes:
        tuyen_name = f"{r.tuyen.tinh_nhan} - {r.tuyen.tinh_giao}"
        unique_routes.add(tuyen_name)
        loai_xe_name = f"{int(r.tai_trong_tan)} tấn" if r.tai_trong_tan.is_integer() else f"{r.tai_trong_tan} tấn"
        unique_types.add(loai_xe_name)
        if r.loai_thung:
            unique_types.add(r.loai_thung)
            
        gcs = GiaCoSo.objects.filter(tuyen=r.tuyen).first()
        if gcs:
            r.gia_co_so_val = gcs.gia_co_so
            r.hieu_luc = f"{gcs.ngay_ap_dung.strftime('%d/%m/%Y')} -> 31/12/2024"

            r.trong_gia = r.muc_gia_chap_nhan <= gcs.gia_co_so
        else:
            r.gia_co_so_val = None
            r.hieu_luc = "-"

            r.trong_gia = False
            
    can_edit = False
    if request.user.is_authenticated:
        can_edit = request.user.is_superuser or partner.nguoi_quan_ly == request.user.username
            
    today = timezone.now().date()
            
    return render(request, 'core/partner_detail.html', {
        'partner': partner, 
        'routes': routes,
        'unique_routes': list(unique_routes),
        'unique_types': list(unique_types),
        'can_edit': can_edit,
        'today': today,
    })

@login_required
def partner_update(request, pk):
    if request.user.is_superuser:
        partner = get_object_or_404(DoiTac, pk=pk)
    else:
        partner = get_object_or_404(DoiTac, pk=pk, nguoi_quan_ly=request.user.username)
        
    routes = partner.bao_gia.all()
    if request.method == 'POST':
        form = DoiTacForm(request.POST, instance=partner, user=request.user)
        if form.is_valid():
            form.save()
            process_routes(request, partner)
            messages.success(request, f'Đã cập nhật đối tác "{partner.ten_nha_xe}" thành công!')
            return redirect('partner_detail', pk=partner.pk)
        else:
            messages.error(request, 'Có lỗi trong form. Vui lòng kiểm tra lại.')
            
            # Giữ lại dữ liệu tuyến mới đã nhập
            tinh_nhans = request.POST.getlist('tinh_nhan[]')
            huyen_nhans = request.POST.getlist('huyen_nhan[]')
            tinh_giaos = request.POST.getlist('tinh_giao[]')
            huyen_giaos = request.POST.getlist('huyen_giao[]')
            tai_trongs = request.POST.getlist('tai_trong_tan[]')
            so_khois = request.POST.getlist('so_khoi[]')
            kich_thuocs = request.POST.getlist('kich_thuoc[]')
            loai_thungs = request.POST.getlist('loai_thung[]')
            muc_gias = request.POST.getlist('muc_gia_chap_nhan[]')
            ghep_hangs = request.POST.getlist('co_ghep_hang_khong[]')
            qua_tais = request.POST.getlist('co_chiu_qua_tai_khong[]')
            nhieu_diems = request.POST.getlist('co_di_nhieu_diem_khong[]')
            chieu_dis = request.POST.getlist('di_1_hay_2_chieu[]')
            bg_ids = request.POST.getlist('bg_id[]')
            
            new_routes_data = []
            for i in range(len(tinh_nhans)):
                # Tất cả tinh_nhans đều là từ form thêm tuyến mới (do tuyến cũ không dùng input tinh_nhan[])
                new_routes_data.append({
                    'tinh_nhan': tinh_nhans[i],
                    'huyen_nhan': huyen_nhans[i] if i < len(huyen_nhans) else '',
                    'tinh_giao': tinh_giaos[i] if i < len(tinh_giaos) else '',
                    'huyen_giao': huyen_giaos[i] if i < len(huyen_giaos) else '',
                    'tai_trong_tan': tai_trongs[i] if i < len(tai_trongs) else '',
                    'so_khoi': so_khois[i] if i < len(so_khois) else '',
                    'kich_thuoc': kich_thuocs[i] if i < len(kich_thuocs) else '',
                    'loai_thung': loai_thungs[i] if i < len(loai_thungs) else '',
                    'muc_gia_chap_nhan': muc_gias[i] if i < len(muc_gias) else '',
                    'co_ghep_hang_khong': ghep_hangs[i] if i < len(ghep_hangs) else '0',
                    'co_chiu_qua_tai_khong': qua_tais[i] if i < len(qua_tais) else '0',
                    'co_di_nhieu_diem_khong': nhieu_diems[i] if i < len(nhieu_diems) else '0',
                    'di_1_hay_2_chieu': chieu_dis[i] if i < len(chieu_dis) else '1'
                })
            
            for r in routes:
                gcs = GiaCoSo.objects.filter(tuyen=r.tuyen).first()
                if gcs:
                    r.gia_co_so_val = gcs.gia_co_so
                    r.trong_gia = r.muc_gia_chap_nhan <= gcs.gia_co_so
                else:
                    r.gia_co_so_val = None
                    r.trong_gia = False
                    
            today = timezone.now().date()
            import json
            return render(request, 'core/partner_form.html', {'form': form, 'title': 'Chỉnh Sửa Đối Tác', 'partner': partner, 'routes': routes, 'today': today, 'new_routes_data_json': json.dumps(new_routes_data)})

    else:
        form = DoiTacForm(instance=partner, user=request.user)
        
    for r in routes:
        gcs = GiaCoSo.objects.filter(tuyen=r.tuyen).first()
        if gcs:
            r.gia_co_so_val = gcs.gia_co_so
            r.trong_gia = r.muc_gia_chap_nhan <= gcs.gia_co_so
        else:
            r.gia_co_so_val = None
            r.trong_gia = False
            
    today = timezone.now().date()
    return render(request, 'core/partner_form.html', {'form': form, 'title': 'Chỉnh Sửa Đối Tác', 'partner': partner, 'routes': routes, 'today': today})

@login_required
def partner_add_route(request, pk):
    if request.method != 'POST':
        return redirect('partner_detail', pk=pk)
        
    if request.user.is_superuser:
        partner = get_object_or_404(DoiTac, pk=pk)
    else:
        partner = get_object_or_404(DoiTac, pk=pk, nguoi_quan_ly=request.user.username)
        
    tinh_nhan_list = request.POST.getlist('tinh_nhan[]')
    huyen_nhan_list = request.POST.getlist('huyen_nhan[]')
    tinh_giao_list = request.POST.getlist('tinh_giao[]')
    huyen_giao_list = request.POST.getlist('huyen_giao[]')
    tai_trong_list = request.POST.getlist('tai_trong_tan[]')
    kich_thuoc_list = request.POST.getlist('kich_thuoc[]')
    loai_thung_list = request.POST.getlist('loai_thung[]')
    muc_gia_list = request.POST.getlist('muc_gia_chap_nhan[]')
    ghep_hang_list = request.POST.getlist('co_ghep_hang_khong[]')
    qua_tai_list = request.POST.getlist('co_chiu_qua_tai_khong[]')
    nhieu_diem_list = request.POST.getlist('co_di_nhieu_diem_khong[]')
    chieu_di_list = request.POST.getlist('di_1_hay_2_chieu[]')
    
    count = 0
    for i in range(len(tinh_nhan_list)):
        t_n = tinh_nhan_list[i]
        h_n = huyen_nhan_list[i] if i < len(huyen_nhan_list) else ''
        t_g = tinh_giao_list[i] if i < len(tinh_giao_list) else ''
        h_g = huyen_giao_list[i] if i < len(huyen_giao_list) else ''
        
        if not t_n or not t_g:
            continue
            
        try:
            tai_trong = float(tai_trong_list[i]) if i < len(tai_trong_list) and tai_trong_list[i] else 0.0
        except ValueError:
            tai_trong = 0.0
        muc_gia = parse_vn_number(muc_gia_list[i]) if i < len(muc_gia_list) and muc_gia_list[i] else 0.0
        ngay_bat_dau = ngay_bat_dau_list[i] if i < len(ngay_bat_dau_list) and ngay_bat_dau_list[i] else timezone.now().date()
        chieu_di = int(chieu_di_list[i]) if i < len(chieu_di_list) and chieu_di_list[i] else 1
            
        kich_thuoc = kich_thuoc_list[i] if i < len(kich_thuoc_list) else ''
        loai_thung = loai_thung_list[i] if i < len(loai_thung_list) else ''
        ghep_hang = True if (i < len(ghep_hang_list) and ghep_hang_list[i] == '1') else False
        qua_tai = True if (i < len(qua_tai_list) and qua_tai_list[i] == '1') else False
        nhieu_diem = True if (i < len(nhieu_diem_list) and nhieu_diem_list[i] == '1') else False
        
        tuyen_queryset = TuyenXe.objects.filter(tinh_nhan=t_n, huyen_nhan=h_n, tinh_giao=t_g, huyen_giao=h_g)
        if tuyen_queryset.exists():
            tuyen = tuyen_queryset.first()
        else:
            import time
            code_t = f"T-{t_n[:2].upper()}-{t_g[:2].upper()}-{int(time.time()*1000) % 100000}"
            tuyen = TuyenXe.objects.create(
                ma_tuyen=code_t,
                tinh_nhan=t_n,
                huyen_nhan=h_n,
                tinh_giao=t_g,
                huyen_giao=h_g
            )
            
        BaoGiaThongTinXe.objects.create(
            doi_tac=partner,
            tuyen=tuyen,
            tai_trong_tan=tai_trong,
            kich_thuoc=kich_thuoc,
            loai_thung=loai_thung,
            muc_gia_chap_nhan=muc_gia,
            
            co_ghep_hang_khong=ghep_hang,
            co_chiu_qua_tai_khong=qua_tai,
            co_di_nhieu_diem_khong=nhieu_diem,
            di_1_hay_2_chieu=chieu_di
        )
        count += 1
        
    if count > 0:
        messages.success(request, f'Đã thêm {count} tuyến thành công!')
    return redirect('partner_detail', pk=pk)

from django.http import JsonResponse
import json

@login_required
def api_route_detail(request, pk):
    try:
        route = BaoGiaThongTinXe.objects.get(pk=pk)
        
        can_edit = False
        if request.user.is_authenticated:
            can_edit = request.user.is_superuser or route.doi_tac.nguoi_quan_ly == request.user.username
            
        history = LichSuBaoGia.objects.filter(bao_gia_ref=route).order_by('-ngay_doi_gia')
        history_data = []
        for h in history:
            history_data.append({
                'ngay_doi_gia': h.ngay_doi_gia.strftime('%d/%m/%Y %H:%M'),
                'gia_cu': h.gia_cu,
                'gia_moi': h.gia_moi,
                'ngay_bat_dau': h.ngay_bat_dau.strftime('%d/%m/%Y') if h.ngay_bat_dau else '-',
                'ngay_ket_thuc': h.ngay_ket_thuc.strftime('%d/%m/%Y') if h.ngay_ket_thuc else '-',
                'gia_xang': h.gia_xang or '-',
                'ly_do_doi': h.ly_do_doi or '-',
                'nguoi_doi': h.nguoi_doi or '-'
            })
            
        return JsonResponse({
            'success': True,
            'id': route.id,
            'ma_tuyen': route.tuyen.ma_tuyen,
            'doi_tac_id': route.doi_tac.id,
            'ma_doi_tac': route.doi_tac.ma_nha_xe or '-',
            'ten_doi_tac': route.doi_tac.ten_nha_xe,
            'sdt_doi_tac': route.doi_tac.so_dien_thoai or 'Không có',
            'tinh_nhan': route.tuyen.tinh_nhan,
            'huyen_nhan': route.tuyen.huyen_nhan,
            'tinh_giao': route.tuyen.tinh_giao,
            'huyen_giao': route.tuyen.huyen_giao,
            'tai_trong_tan': route.tai_trong_tan,
            'so_khoi': route.so_khoi or '',
            'kich_thuoc': route.kich_thuoc or '',
            'loai_thung': route.loai_thung or '',
            'muc_gia_chap_nhan': route.muc_gia_chap_nhan,
            'ngay_bat_dau': route.ngay_bat_dau.strftime('%Y-%m-%d') if route.ngay_bat_dau else '',
            'co_ghep_hang_khong': route.co_ghep_hang_khong,
            'co_chiu_qua_tai_khong': route.co_chiu_qua_tai_khong,
            'co_di_nhieu_diem_khong': route.co_di_nhieu_diem_khong,
            'di_1_hay_2_chieu': route.di_1_hay_2_chieu,
            'history': history_data,
            'can_edit': can_edit
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_route_update(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        data = json.loads(request.body)
        if request.user.is_superuser:
            route = BaoGiaThongTinXe.objects.get(pk=pk)
        else:
            route = BaoGiaThongTinXe.objects.get(pk=pk, doi_tac__nguoi_quan_ly=request.user.username)
            
        # Optional fields updates
        if 'tai_trong_tan' in data: route.tai_trong_tan = parse_vn_number(data['tai_trong_tan'])
        if 'so_khoi' in data: route.so_khoi = parse_vn_number(data['so_khoi']) if data['so_khoi'] else None
        if 'kich_thuoc' in data: route.kich_thuoc = data['kich_thuoc']
        if 'loai_thung' in data: route.loai_thung = data['loai_thung']
        if 'co_ghep_hang_khong' in data: route.co_ghep_hang_khong = bool(data['co_ghep_hang_khong'])
        if 'co_chiu_qua_tai_khong' in data: route.co_chiu_qua_tai_khong = bool(data['co_chiu_qua_tai_khong'])
        if 'co_di_nhieu_diem_khong' in data: route.co_di_nhieu_diem_khong = bool(data['co_di_nhieu_diem_khong'])
        if 'di_1_hay_2_chieu' in data: route.di_1_hay_2_chieu = int(data['di_1_hay_2_chieu'])
        
        # Price update logic
        # Price update logic
        if 'muc_gia_chap_nhan' in data:
            new_price = parse_vn_number(data['muc_gia_chap_nhan'])
            route._gia_xang = parse_vn_number(data.get('gia_xang')) if data.get('gia_xang') else None
            route._ly_do_doi = data.get('ly_do_doi') or 'Cập nhật giá'
            route._nguoi_doi = request.user.username
            route.muc_gia_chap_nhan = new_price
            route._force_history = True
        route.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_delete_route(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        if request.user.is_superuser:
            route = BaoGiaThongTinXe.objects.get(pk=pk)
        else:
            route = BaoGiaThongTinXe.objects.get(pk=pk, doi_tac__nguoi_quan_ly=request.user.username)
            
        route.is_deleted = True
        route.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
from datetime import date
from django.db.models import Q
from .models import LichSuGiaCoSo

@login_required
def base_price_list(request):
    tinh_nhan_qs = TuyenXe.objects.values_list('tinh_nhan', flat=True).distinct()
    tinh_giao_qs = TuyenXe.objects.values_list('tinh_giao', flat=True).distinct()
    tinh_thanh = sorted(list(set(tinh_nhan_qs) | set(tinh_giao_qs)))
    loai_xes = sorted(list(GiaCoSo.objects.values_list('loai_xe', flat=True).distinct()))
    return render(request, 'core/base_price_list.html', {
        'tinh_thanh': tinh_thanh,
        'loai_xes': loai_xes,
    })

@login_required
def api_get_base_prices(request):
    try:
        tinh_nhan = request.GET.get('tinh_nhan', '')
        huyen_nhan = request.GET.get('huyen_nhan', '')
        tinh_giao = request.GET.get('tinh_giao', '')
        huyen_giao = request.GET.get('huyen_giao', '')
        so_khoi = request.GET.get('so_khoi', '')
        loai_xe = request.GET.get('loai_xe', '')
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        huyen_nhan = clean_huyen(huyen_nhan)
        huyen_giao = clean_huyen(huyen_giao)
            
        queryset = GiaCoSo.objects.select_related('tuyen').all().order_by('-id')
        
        if tinh_nhan: queryset = queryset.filter(tuyen__tinh_nhan__icontains=tinh_nhan)
        if huyen_nhan: queryset = queryset.filter(tuyen__huyen_nhan__icontains=huyen_nhan)
        if tinh_giao: queryset = queryset.filter(tuyen__tinh_giao__icontains=tinh_giao)
        if huyen_giao: queryset = queryset.filter(tuyen__huyen_giao__icontains=huyen_giao)
        if so_khoi: queryset = queryset.filter(so_khoi=so_khoi)
        if loai_xe: queryset = queryset.filter(loai_xe__icontains=loai_xe)
        
        data = []
        for g in queryset:
            data.append({
                'id': g.id,
                'tinh_nhan': g.tuyen.tinh_nhan,
                'huyen_nhan': g.tuyen.huyen_nhan,
                'tinh_giao': g.tuyen.tinh_giao,
                'huyen_giao': g.tuyen.huyen_giao,
                'so_khoi': g.so_khoi or '-',
                'loai_xe': g.loai_xe,
                'gia_co_so': g.gia_co_so,
                'ngay_ap_dung': g.ngay_ap_dung.strftime('%d/%m/%Y') if g.ngay_ap_dung else '-'
            })
            
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_export_base_prices_info(request):
    try:
        tinh_nhan = request.GET.get('tinh_nhan', '')
        huyen_nhan = request.GET.get('huyen_nhan', '')
        tinh_giao = request.GET.get('tinh_giao', '')
        huyen_giao = request.GET.get('huyen_giao', '')
        so_khoi = request.GET.get('so_khoi', '')
        loai_xe = request.GET.get('loai_xe', '')
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        huyen_nhan = clean_huyen(huyen_nhan)
        huyen_giao = clean_huyen(huyen_giao)
            
        queryset = GiaCoSo.objects.select_related('tuyen').all()
        
        if tinh_nhan: queryset = queryset.filter(tuyen__tinh_nhan__icontains=tinh_nhan)
        if huyen_nhan: queryset = queryset.filter(tuyen__huyen_nhan__icontains=huyen_nhan)
        if tinh_giao: queryset = queryset.filter(tuyen__tinh_giao__icontains=tinh_giao)
        if huyen_giao: queryset = queryset.filter(tuyen__huyen_giao__icontains=huyen_giao)
        if so_khoi: queryset = queryset.filter(so_khoi=so_khoi)
        if loai_xe: queryset = queryset.filter(loai_xe__icontains=loai_xe)
        
        ids_str = request.GET.get('ids', '')
        if ids_str:
            ids_list = [int(id.strip()) for id in ids_str.split(',') if id.strip().isdigit()]
            if ids_list:
                queryset = queryset.filter(id__in=ids_list)
        
        routes = set()
        for g in queryset:
            routes.add((g.tuyen.huyen_nhan, g.tuyen.tinh_nhan, g.tuyen.tinh_giao, g.tuyen.huyen_giao))
            
        return JsonResponse({
            'success': True,
            'total_routes': len(routes),
            'total_prices': queryset.count(),
            'format': 'Excel (.xlsx)'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def autofit_columns(ws, skip_rows=2):
    import openpyxl
    for idx, col in enumerate(ws.columns):
        max_length = 0
        column = openpyxl.utils.get_column_letter(idx + 1)
        for cell in col:
            if cell.row <= skip_rows: continue
            # If the cell is merged, its length shouldn't dictate the column width
            if type(cell).__name__ == 'MergedCell': continue
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2) * 1.1
        if adjusted_width < 12: adjusted_width = 12
        if adjusted_width > 40: adjusted_width = 40
        ws.column_dimensions[column].width = adjusted_width

@login_required
def api_export_base_prices_excel(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        tinh_nhan = request.GET.get('tinh_nhan', '')
        huyen_nhan = request.GET.get('huyen_nhan', '')
        tinh_giao = request.GET.get('tinh_giao', '')
        huyen_giao = request.GET.get('huyen_giao', '')
        so_khoi = request.GET.get('so_khoi', '')
        loai_xe = request.GET.get('loai_xe', '')
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        huyen_nhan = clean_huyen(huyen_nhan)
        huyen_giao = clean_huyen(huyen_giao)
            
        queryset = GiaCoSo.objects.select_related('tuyen').all()
        
        if tinh_nhan: queryset = queryset.filter(tuyen__tinh_nhan__icontains=tinh_nhan)
        if huyen_nhan: queryset = queryset.filter(tuyen__huyen_nhan__icontains=huyen_nhan)
        if tinh_giao: queryset = queryset.filter(tuyen__tinh_giao__icontains=tinh_giao)
        if huyen_giao: queryset = queryset.filter(tuyen__huyen_giao__icontains=huyen_giao)
        if so_khoi: queryset = queryset.filter(so_khoi=so_khoi)
        if loai_xe: queryset = queryset.filter(loai_xe__icontains=loai_xe)
        
        ids_str = request.GET.get('ids', '')
        if ids_str:
            ids_list = [int(id.strip()) for id in ids_str.split(',') if id.strip().isdigit()]
            if ids_list:
                queryset = queryset.filter(id__in=ids_list)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bảng giá cơ sở"
        
        # Sheet "Data" for dropdowns
        ws_data = wb.create_sheet(title="Data")
        ws_data.sheet_state = 'hidden'

        import urllib.request
        import json
        from django.core.cache import cache
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter

        provinces_data = cache.get('provinces_data_full')
        if not provinces_data:
            try:
                req = urllib.request.Request('https://provinces.open-api.vn/api/?depth=2', headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=5) as response:
                    provinces_data = json.loads(response.read().decode())
                cache.set('provinces_data_full', provinces_data, 86400)
            except Exception as e:
                provinces_data = []

        def clean_name(s):
            return s.replace(" ", "").replace("-", "").replace(".", "")

        province_names = []
        col_idx = 2
        for p in provinces_data:
            p_name = p['name'].replace("Tỉnh ", "").replace("Thành phố ", "")
            province_names.append(p_name)
            
            d_names = [d['name'] for d in p['districts']]
            ws_data.cell(row=1, column=col_idx, value=p_name)
            for r_idx, d_name in enumerate(d_names, start=2):
                ws_data.cell(row=r_idx, column=col_idx, value=d_name)
            
            c_letter = get_column_letter(col_idx)
            ref = f"Data!${c_letter}$2:${c_letter}${len(d_names)+1}"
            c_name = clean_name(p_name)
            try:
                wb.create_named_range(c_name, None, ref)
            except Exception as e:
                pass
            
            col_idx += 1

        ws_data.cell(row=1, column=1, value="Provinces")
        for r_idx, p_name in enumerate(province_names, start=2):
            ws_data.cell(row=r_idx, column=1, value=p_name)
        
        wb.create_named_range("AllProvinces", None, f"Data!$A$2:$A${len(province_names)+1}")
        
        ws.append(["CÔNG TY CỔ PHẦN NHẤT PHONG VẬN"])
        ws.append(["BẢNG GIÁ CƠ SỞ"])
        from datetime import datetime
        current_date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws.append([f"Ngày xuất: {current_date_str}"])
        ws.append([])
        
        ws.merge_cells("A1:M1")
        ws.merge_cells("A2:M2")
        ws.merge_cells("A3:M3")
        
        title_font = Font(size=16, bold=True, color="FF0000")
        subtitle_font = Font(size=14, bold=True, color="0000FF")
        date_font = Font(size=11, italic=True)
        align_center = Alignment(horizontal="center", vertical="center")
        
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = align_center
        ws["A3"].font = date_font
        ws["A3"].alignment = align_center
        
        row5 = ["Tỉnh nhận", "Huyện nhận", "Tỉnh giao", "Huyện giao", "Loại xe", "1.25T", "2.5T", "3.5T", "5T", "7T", "8T", "9T", "15T", "LTL"]
        
        from core.models import CauHinhLoaiXe
        default_max = {"1.25T": 9, "2.5T": 13, "3.5T": 18, "5T": 23, "7T": 35, "8T": 54, "9T": 40, "15T": 54}
        row6 = ["", "", "", "", "Số khối"]
        for lx in ["1.25T", "2.5T", "3.5T", "5T", "7T", "8T", "9T", "15T"]:
            ch = CauHinhLoaiXe.objects.filter(loai_xe=lx).first()
            if ch and ch.khoi_den is not None:
                row6.append(f"{ch.khoi_den:g}")
            else:
                row6.append(str(default_max[lx]))
        row6.append("LTL")
        ws.append(row5)
        ws.append(row6)
        
        ws.merge_cells("A5:A6")
        ws.merge_cells("B5:B6")
        ws.merge_cells("C5:C6")
        ws.merge_cells("D5:D6")
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for r in [5, 6]:
            for c in range(1, 15):
                cell = ws.cell(row=r, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
                
        # Add Data Validations
        if province_names:
            dv_prov = DataValidation(type="list", formula1="=AllProvinces", allow_blank=True)
            ws.add_data_validation(dv_prov)
            dv_prov.add('A7:A1000')
            dv_prov.add('C7:C1000')

            dv_dist_nhan = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(A7, " ", ""), "-", ""), ".", ""))', allow_blank=True)
            ws.add_data_validation(dv_dist_nhan)
            dv_dist_nhan.add('B7:B1000')

            dv_dist_giao = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(C7, " ", ""), "-", ""), ".", ""))', allow_blank=True)
            ws.add_data_validation(dv_dist_giao)
            dv_dist_giao.add('D7:D1000')
        
        grouped_data = {}
        for g in queryset:
            route_key = (g.tuyen.tinh_nhan or "", g.tuyen.huyen_nhan or "", g.tuyen.tinh_giao or "", g.tuyen.huyen_giao or "")
            if route_key not in grouped_data:
                grouped_data[route_key] = {}
            grouped_data[route_key][g.loai_xe.strip()] = g.gia_co_so
            
        lx_cols = {
            "1.25T": 6, "2.5T": 7, "3.5T": 8, "5T": 9,
            "7T": 10, "8T": 11, "9T": 12, "15T": 13, "LTL": 14
        }
        
        sorted_routes = sorted(grouped_data.keys(), key=lambda x: (x[0], x[1], x[2], x[3]))
        
        fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_even = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
        
        row_idx = 7
        for route_key in sorted_routes:
            prices = grouped_data[route_key]
            row_data = [route_key[0], route_key[1], route_key[2], route_key[3], ""] + [""] * 9
            for lx, price in prices.items():
                col_idx = lx_cols.get(lx)
                if col_idx:
                    row_data[col_idx - 1] = price
            ws.append(row_data)
            
            current_fill = fill_even if row_idx % 2 == 0 else fill_odd
            for c in range(1, 15):
                cell = ws.cell(row=row_idx, column=c)
                cell.fill = current_fill
                cell.border = thin_border
                if c >= 6 and cell.value != "":
                    cell.number_format = '#,##0'
            row_idx += 1
                    
        for col_letter in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions['E'].width = 12
        for col_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col_letter].width = 16
            
        # Format empty rows for prices (columns F to M)
        for row in range(6, 1001):
            for col in range(6, 14):
                c = ws.cell(row=row, column=col)
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right', vertical='center')
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="gia_co_so.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error exporting to Excel: {str(e)}", status=500)

@login_required
def api_base_price_detail(request, pk):
    try:
        g = GiaCoSo.objects.get(pk=pk)
        history = LichSuGiaCoSo.objects.filter(gia_co_so_ref=g).order_by('-ngay_doi')
        
        history_data = []
        for h in history:
            history_data.append({
                'ngay_doi': h.ngay_doi.strftime('%d/%m/%Y'),
                'gia_cu': h.gia_cu,
                'gia_moi': h.gia_moi,
                'gia_xang': h.gia_xang or '-',
                'ly_do_doi': h.ly_do_doi or '-'
            })
            
        return JsonResponse({
            'success': True,
            'id': g.id,
            'tuyen': f"{g.tuyen.tinh_nhan} -> {g.tuyen.tinh_giao}",
            'loai_xe': g.loai_xe,
            'so_khoi': g.so_khoi or '',
            'gia_co_so': g.gia_co_so,
            'ngay_ap_dung': g.ngay_ap_dung.strftime('%Y-%m-%d') if g.ngay_ap_dung else '',
            'history': history_data
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_save_base_price(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Chỉ admin mới có quyền thêm giá cơ sở'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        data = json.loads(request.body)
        
        tinh_nhan = data.get('tinh_nhan', '').strip()
        huyen_nhan = data.get('huyen_nhan', '').strip()
        tinh_giao = data.get('tinh_giao', '').strip()
        huyen_giao = data.get('huyen_giao', '').strip()
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        huyen_nhan = clean_huyen(huyen_nhan)
        huyen_giao = clean_huyen(huyen_giao)
        
        loai_xe = data.get('loai_xe')
        so_khoi = data.get('so_khoi') if data.get('so_khoi') else None
        gia_co_so = parse_vn_number(data.get('gia_co_so', 0))
        ngay_ap_dung_str = data.get('ngay_ap_dung')
        ly_do_doi = data.get('ly_do_doi')
        if not ly_do_doi or ly_do_doi.strip() == '':
            ly_do_doi = 'Tạo mới/Cập nhật giá'
        
        gia_xang = data.get('gia_xang')
        gia_xang = parse_vn_number(gia_xang) if gia_xang else None
        
        ngay_ap_dung = date.today()
        if ngay_ap_dung_str:
            from datetime import datetime
            ngay_ap_dung = datetime.strptime(ngay_ap_dung_str, '%Y-%m-%d').date()
            
        if not tinh_nhan or not tinh_giao or not loai_xe:
            return JsonResponse({'success': False, 'error': 'Thiếu thông tin bắt buộc'})
            
        tuyen, created = TuyenXe.objects.get_or_create(
            tinh_nhan=tinh_nhan,
            huyen_nhan=huyen_nhan,
            tinh_giao=tinh_giao,
            huyen_giao=huyen_giao,
            defaults={'ma_tuyen': f"T_{int(datetime.now().timestamp())}"}
        )
        
        # Check if GiaCoSo exists
        existing_gcs = GiaCoSo.objects.filter(tuyen=tuyen, loai_xe=loai_xe).first()
        if existing_gcs:
            return JsonResponse({
                'success': False, 
                'exists': True, 
                'id': existing_gcs.id, 
                'error': f"Giá cơ sở cho tuyến này với loại xe {loai_xe} đã tồn tại."
            })
            
        g = GiaCoSo(
            tuyen=tuyen,
            loai_xe=loai_xe,
            so_khoi=so_khoi,
            gia_co_so=gia_co_so,
            ngay_ap_dung=ngay_ap_dung
        )
        g._ly_do_doi = ly_do_doi
        g._gia_xang = gia_xang
        g.save()
        
        return JsonResponse({'success': True, 'id': g.id})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_update_base_price(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Chỉ admin mới có quyền chỉnh sửa giá cơ sở'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        data = json.loads(request.body)
        g = GiaCoSo.objects.get(pk=pk)
        
        if 'gia_co_so' in data:
            g.gia_co_so = parse_vn_number(data['gia_co_so'])
            
        if data.get('ngay_ap_dung'):
            from datetime import datetime
            g.ngay_ap_dung = datetime.strptime(data['ngay_ap_dung'], '%Y-%m-%d').date()
            
        ly_do_doi = data.get('ly_do_doi')
        if not ly_do_doi or ly_do_doi.strip() == '':
            return JsonResponse({'success': False, 'error': 'Lý do cập nhật là bắt buộc'})
            
        g._ly_do_doi = ly_do_doi
        g.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def search_prices_view(request):
    return render(request, 'core/search_prices.html')

@login_required
def api_search_prices(request):
    try:
        tinh_nhan = request.GET.get('tinh_nhan', '').strip()
        huyen_nhan = request.GET.get('huyen_nhan', '').strip()
        tinh_giao = request.GET.get('tinh_giao', '').strip()
        huyen_giao = request.GET.get('huyen_giao', '').strip()
        loai_xe = request.GET.get('loai_xe', '').strip()
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
        
        
        # Build base quote query
        if request.user.is_superuser:
            bg_qs = BaoGiaThongTinXe.objects.select_related('doi_tac', 'tuyen').all()
        else:
            bg_qs = BaoGiaThongTinXe.objects.select_related('doi_tac', 'tuyen').filter(doi_tac__nguoi_quan_ly=request.user.username)
        
        if tinh_nhan:
            bg_qs = bg_qs.filter(tuyen__tinh_nhan=tinh_nhan)
        if huyen_nhan:
            huyen_nhan = clean_huyen(huyen_nhan)
            bg_qs = bg_qs.filter(tuyen__huyen_nhan=huyen_nhan)
        if tinh_giao:
            bg_qs = bg_qs.filter(tuyen__tinh_giao=tinh_giao)
        if huyen_giao:
            huyen_giao = clean_huyen(huyen_giao)
            bg_qs = bg_qs.filter(tuyen__huyen_giao=huyen_giao)
        if loai_xe:
            try:
                loai_xe_val = float(loai_xe.lower().replace('t', '').strip())
                bg_qs = bg_qs.filter(tai_trong_tan=loai_xe_val)
            except ValueError:
                pass
                
        so_khoi = request.GET.get('so_khoi', '').strip()
        if so_khoi:
            bg_qs = bg_qs.filter(so_khoi=so_khoi)
                
        bg_qs = bg_qs.order_by('muc_gia_chap_nhan')
        
        # Pre-fetch Base Prices for quick lookup
        gcs_dict = {}
        for g in GiaCoSo.objects.select_related('tuyen').all():
            key = (g.tuyen.id, g.loai_xe)
            gcs_dict[key] = g.gia_co_so

        def get_lx_str(tt):
            if tt.is_integer():
                return f"{int(tt)}T"
            return f"{tt}T"

        results = []
        is_exact_search = bool(tinh_nhan and tinh_giao and (loai_xe or so_khoi))
        exact_base_price = None

        for bg in bg_qs:
            lx_str = get_lx_str(bg.tai_trong_tan)
            key = (bg.tuyen.id, lx_str)
            g_val = gcs_dict.get(key)
            
            # The user requested to show all results regardless of base price
            # if g_val is not None:
            #     if bg.muc_gia_chap_nhan > g_val:
            #         continue
                
            # For exact search info box
            if is_exact_search and g_val is not None and not exact_base_price:
                # We just find the first match to display if they did an exact search
                g_obj = GiaCoSo.objects.filter(tuyen=bg.tuyen, loai_xe=lx_str).first()
                if g_obj:
                    exact_base_price = {
                        'gia_co_so': g_obj.gia_co_so,
                        'so_khoi': g_obj.so_khoi or '-',
                        'ngay_ap_dung': g_obj.ngay_ap_dung.strftime('%d/%m/%Y') if g_obj.ngay_ap_dung else '-'
                    }

            results.append({
                'bg_id': bg.id,
                'doi_tac_id': bg.doi_tac.id,
                'doi_tac_ten': bg.doi_tac.ten_nha_xe,
                'doi_tac_sdt': bg.doi_tac.so_dien_thoai or 'Không có',
                'tuyen': f"{bg.tuyen.huyen_nhan + ', ' if bg.tuyen.huyen_nhan else ''}{bg.tuyen.tinh_nhan} ➔ {bg.tuyen.huyen_giao + ', ' if bg.tuyen.huyen_giao else ''}{bg.tuyen.tinh_giao}",
                'loai_xe': lx_str,
                'so_khoi': bg.so_khoi or '-',
                'kich_thuoc': bg.kich_thuoc or '-',
                'loai_thung': bg.loai_thung or '-',
                'muc_gia': bg.muc_gia_chap_nhan,
                'ghep_hang': 'Có' if bg.co_ghep_hang_khong else 'Không',
                'chieu_di': f"{bg.di_1_hay_2_chieu} chiều",
                'gia_co_so': g_val,
            })
            
        return JsonResponse({'success': True, 'is_exact_search': is_exact_search, 'base_price': exact_base_price, 'results': results})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def proposal_list_view(request):
    if request.user.is_superuser:
        proposals = DeXuatGiaCoSo.objects.all().order_by('-ngay_tao')
    else:
        proposals = DeXuatGiaCoSo.objects.filter(nguoi_de_xuat=request.user.username).order_by('-ngay_tao')
        DeXuatGiaCoSo.objects.filter(nguoi_de_xuat=request.user.username, is_read=False).exclude(trang_thai='ChoDuyet').update(is_read=True)
    
    tinh_nhan_qs = TuyenXe.objects.values_list('tinh_nhan', flat=True).distinct()
    tinh_giao_qs = TuyenXe.objects.values_list('tinh_giao', flat=True).distinct()
    tinh_thanh = sorted(list(set(tinh_nhan_qs) | set(tinh_giao_qs)))
    loai_xes = sorted(list(GiaCoSo.objects.values_list('loai_xe', flat=True).distinct()))
    
    return render(request, 'core/proposal_list.html', {
        'proposals': proposals,
        'tinh_thanh': tinh_thanh,
        'loai_xes': loai_xes
    })

@login_required
def api_create_proposal(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        data = json.loads(request.body)
        tinh_nhan = data.get('tinh_nhan', '').strip()
        huyen_nhan = data.get('huyen_nhan', '').strip()
        tinh_giao = data.get('tinh_giao', '').strip()
        huyen_giao = data.get('huyen_giao', '').strip()
        loai_xe = data.get('loai_xe', '').strip()
        so_khoi = data.get('so_khoi', '').strip()
        gia_de_xuat = data.get('gia_de_xuat')
        ly_do = data.get('ly_do')
        
        prefixes_tinh = ["Tỉnh ", "Thành phố "]
        for p in prefixes_tinh:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        huyen_nhan = clean_huyen(huyen_nhan)
        huyen_giao = clean_huyen(huyen_giao)
        
        if not (tinh_nhan and tinh_giao and loai_xe and gia_de_xuat and ly_do):
            return JsonResponse({'success': False, 'error': 'Thiếu thông tin bắt buộc'})
            
        tuyen, created = TuyenXe.objects.get_or_create(
            tinh_nhan=tinh_nhan, huyen_nhan=huyen_nhan, tinh_giao=tinh_giao, huyen_giao=huyen_giao,
            defaults={'ma_tuyen': f"T-DX-{int(time.time())}"}
        )
        
        gcs = GiaCoSo.objects.filter(tuyen=tuyen, loai_xe=loai_xe).first()
        gia_hien_tai = gcs.gia_co_so if gcs else None
        
        gia_de_xuat_clean = str(gia_de_xuat).replace(',', '').replace(' ', '')
        
        DeXuatGiaCoSo.objects.create(
            tuyen=tuyen, loai_xe=loai_xe, so_khoi=so_khoi,
            gia_hien_tai=gia_hien_tai, gia_de_xuat=parse_vn_number(gia_de_xuat_clean),
            ly_do_de_xuat=ly_do, nguoi_de_xuat=request.user.username
        )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_approve_proposal(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
        
    try:
        proposal = DeXuatGiaCoSo.objects.get(pk=pk)
        if proposal.trang_thai != 'ChoDuyet':
            return JsonResponse({'success': False, 'error': 'Đề xuất này đã được xử lý'})
            
        proposal.trang_thai = 'DaDuyet'
        proposal.is_read = False
        proposal.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_reject_proposal(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
        
    try:
        data = json.loads(request.body)
        proposal = DeXuatGiaCoSo.objects.get(pk=pk)
        if proposal.trang_thai != 'ChoDuyet':
            return JsonResponse({'success': False, 'error': 'Đề xuất này đã được xử lý'})
            
        proposal.trang_thai = 'TuChoi'
        proposal.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_get_base_price(request):
    try:
        tinh_nhan = request.GET.get('tinh_nhan', '').strip()
        tinh_giao = request.GET.get('tinh_giao', '').strip()
        loai_xe = request.GET.get('loai_xe', '').strip()
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        gcs = GiaCoSo.objects.filter(tuyen__tinh_nhan=tinh_nhan, tuyen__tinh_giao=tinh_giao, loai_xe=loai_xe).first()
        if gcs:
            return JsonResponse({'success': True, 'gia_co_so': gcs.gia_co_so, 'so_khoi': gcs.so_khoi})
        return JsonResponse({'success': True, 'gia_co_so': None})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_delete_partner(request, pk):
    if request.method != 'POST': return JsonResponse({'success': False})
    try:
        if request.user.is_superuser: p = DoiTac.objects.get(pk=pk)
        else: p = DoiTac.objects.get(pk=pk, nguoi_quan_ly=request.user.username)
        p.is_deleted = True; p.save()
        return JsonResponse({'success': True})
    except Exception as e: return JsonResponse({'success': False})

def export_single_partner_excel(partner):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.utils import timezone
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hồ Sơ Đối Tác"
    
    # Fonts & Fills
    title_font = Font(name='Arial', size=16, bold=True, color='FF0000')
    subtitle_font = Font(name='Arial', size=14, bold=True, color='0000FF')
    italic_font = Font(name='Arial', size=11, italic=True)
    
    section_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    section_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    label_font = Font(name='Arial', size=11, bold=True)
    value_font = Font(name='Arial', size=11)
    
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1. Header
    ws.merge_cells('A1:K1')
    c = ws['A1']
    c.value = "CÔNG TY CỔ PHẦN NHẤT PHONG VẬN"
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:K2')
    c = ws['A2']
    c.value = "THÔNG TIN & BẢNG GIÁ ĐỐI TÁC"
    c.font = subtitle_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A3:K3')
    c = ws['A3']
    c.value = f"Ngày xuất: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}"
    c.font = italic_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    
    # 2. THÔNG TIN CHUNG
    ws.merge_cells('A5:K5')
    c = ws['A5']
    c.value = "THÔNG TIN CHUNG"
    c.font = section_font
    c.fill = section_fill
    c.alignment = Alignment(vertical='center')
    
    info_layout = [
        ("Mã nhà xe:", partner.ma_nha_xe or "", "Số điện thoại:", partner.so_dien_thoai or ""),
        ("Tên nhà xe:", partner.ten_nha_xe or "", "Zalo:", partner.ten_zalo or ""),
        ("Địa chỉ:", partner.dia_chi or "", "Người quản lý:", partner.nguoi_quan_ly or ""),
        ("Mã số thuế:", partner.ma_so_thue or "", "Người đại diện:", partner.nguoi_dai_dien or ""),
        ("ĐV Xuất HĐ:", partner.ten_don_vi_xuat_hoa_don or "", "Địa chỉ HĐ:", partner.dia_chi_xuat_hoa_don or ""),
        ("Số tài khoản:", partner.so_tai_khoan or "", "Ngân hàng:", partner.ngan_hang or ""),
        ("Sẵn sàng:", partner.tinh_san_sang or "", "Hợp đồng:", "Có" if partner.co_hop_dong else "Không"),
        ("Hạn TT:", partner.thoi_han_thanh_toan or "", "Chất lượng:", partner.chat_luong or ""),
        ("Phí bốc xếp:", f"{partner.phi_boc_xep:,.0f}" if partner.phi_boc_xep else "", "Ghi chú:", partner.ghi_chu or ""),
        ("Phí rớt điểm:", f"{partner.phi_rot_diem:,.0f}" if partner.phi_rot_diem else "", "", ""),
        ("Phí quá tải:", f"{partner.phi_di_qua_tai:,.0f}" if partner.phi_di_qua_tai else "", "", "")
    ]
    
    row = 6
    for l1, v1, l2, v2 in info_layout:
        ws.cell(row=row, column=1, value=l1).font = label_font
        ws.cell(row=row, column=2, value=v1).font = value_font
        ws.cell(row=row, column=3, value=l2).font = label_font
        ws.cell(row=row, column=4, value=v2).font = value_font
        row += 1
        
    row += 1
    # 3. BẢNG GIÁ CHI TIẾT CÁC TUYẾN
    ws.merge_cells(f'A{row}:K{row}')
    c = ws[f'A{row}']
    c.value = "BẢNG GIÁ CHI TIẾT CÁC TUYẾN"
    c.font = section_font
    c.fill = section_fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    # Headers Row 1
    ws.merge_cells(f'A{row}:A{row+1}')
    c = ws[f'A{row}']
    c.value = "STT"
    
    ws.merge_cells(f'B{row}:E{row}')
    c = ws[f'B{row}']
    c.value = "TUYẾN XE"
    
    ws.merge_cells(f'F{row}:F{row+1}')
    c = ws[f'F{row}']
    c.value = "ĐI NHIỀU ĐIỂM"
    
    ws.merge_cells(f'G{row}:J{row}')
    c = ws[f'G{row}']
    c.value = "NĂNG LỰC XE"
    
    ws.merge_cells(f'K{row}:K{row+1}')
    c = ws[f'K{row}']
    c.value = "MỨC GIÁ CHẤP NHẬN"
    
    # Headers Row 2
    headers_r2 = {
        2: "Tỉnh nhận", 3: "Huyện nhận", 4: "Tỉnh giao", 5: "Huyện giao",
        7: "Tải trọng (Tấn)", 8: "Số khối", 9: "Kích thước", 10: "Loại thùng"
    }
    for col_idx, text in headers_r2.items():
        c = ws.cell(row=row+1, column=col_idx, value=text)
        
    # Apply styles to headers
    for r_idx in range(row, row+2):
        for c_idx in range(1, 12):
            c = ws.cell(row=r_idx, column=c_idx)
            c.font = header_font
            c.fill = section_fill
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
    row += 2
    
    routes = partner.bao_gia.filter(is_deleted=False).order_by('id')
    fill_even = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    fill_odd = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for i, r in enumerate(routes, 1):
        row_values = [
            i, r.tuyen.tinh_nhan, r.tuyen.huyen_nhan, r.tuyen.tinh_giao, r.tuyen.huyen_giao,
            "Có" if r.co_di_nhieu_diem_khong else "Không",
            r.tai_trong_tan, r.so_khoi, r.kich_thuoc, r.loai_thung, r.muc_gia_chap_nhan
        ]
        current_fill = fill_even if i % 2 != 0 else fill_odd
        for col_idx, val in enumerate(row_values, 1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.font = value_font
            c.border = thin_border
            c.fill = current_fill
            if col_idx in [6, 7, 8]: c.alignment = Alignment(horizontal='center')
            elif col_idx == 1: c.alignment = Alignment(horizontal='right')
            elif col_idx == 11: 
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right')
        row += 1
        
    autofit_columns(ws, skip_rows=4)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    import unicodedata
    import re
    safe_name = unicodedata.normalize('NFKD', partner.ten_nha_xe).encode('ASCII', 'ignore').decode('utf-8')
    safe_name = re.sub(r'[^\w\s-]', '', safe_name).strip().replace(' ', '_')
    if not safe_name: safe_name = f"partner_{partner.pk}"
    response['Content-Disposition'] = f'attachment; filename="Ho_so_{safe_name}.xlsx"'
    wb.save(response)
    return response

@login_required
def api_export_partners_excel(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.db.models import Q
        from core.models import DoiTac
        
        search_query = request.GET.get('search', '')
        from django.db.models import Prefetch
        partners = DoiTac.objects.prefetch_related('bao_gia__tuyen').all().order_by('-id')
        
        if search_query:
            partners = partners.filter(
                Q(ten_nha_xe__icontains=search_query) |
                Q(ma_nha_xe__icontains=search_query) |
                Q(so_dien_thoai__icontains=search_query) |
                Q(nguoi_quan_ly__icontains=search_query)
            )
            
        partner_id = request.GET.get('partner_id')
        ids_param = request.GET.get('ids')
        if partner_id:
            p = partners.filter(id=partner_id).first()
            if p:
                return export_single_partner_excel(p)
            else:
                from django.http import HttpResponse
                return HttpResponse("Không tìm thấy đối tác", status=404)
        elif ids_param:
            ids_list = [int(i) for i in ids_param.split(',') if i.isdigit()]
            if ids_list:
                partners = partners.filter(id__in=ids_list)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DanhSachDoiTac"
        
        ws.append(["CÔNG TY CỔ PHẦN NHẤT PHONG VẬN"])
        ws.append(["DANH SÁCH ĐỐI TÁC THUÊ NGOÀI VÀ BẢNG GIÁ"])
        ws.append([])
        
        # Headers level 1
        headers_l1 = [
            "STT", "Mã nhà xe", "Tên nhà xe", "Số điện thoại", "Địa chỉ", "Tên Zalo",
            "THÔNG TIN XUẤT HOÁ ĐƠN", "", "", "", "", "",
            "ĐÁNH GIÁ & THANH TOÁN", "", "", "", "",
            "CÁC LOẠI CHI PHÍ NHÀ XE", "", "",
            "TUYẾN XE", "", "", "",
            "THÔNG TIN XE & NĂNG LỰC", "", "", "", "", "", "", "",
            "MỨC GIÁ CHẤP NHẬN", "NGƯỜI QUẢN LÝ"
        ]
        
        # Headers level 2
        headers_l2 = [
            "STT", "Mã nhà xe", "Tên nhà xe", "Số điện thoại", "Địa chỉ", "Tên Zalo",
            "Tên đơn vị", "Địa chỉ HĐ", "Mã số thuế", "Số tài khoản", "Người đại diện", "Ngân hàng",
            "Thời hạn thanh toán", "Tính sẵn sàng", "Có hợp đồng không", "Chất lượng", "Ghi chú",
            "Phí rớt điểm", "Phí bốc xếp", "Phí đi quá tải",
            "Tỉnh nhận", "Huyện nhận", "Tỉnh giao", "Huyện giao",
            "Số chiều đi", "Đi nhiều điểm", "Ghép hàng", "Chịu quá tải", "Tải trọng (Tấn)", "Số khối", "Kích thước", "Loại thùng",
            "Mức giá (VNĐ)", "Người quản lý"
        ]
        
        ws.append(headers_l1)
        ws.append(headers_l2)
        
        # Merging level 1 headers
        ws.merge_cells("A1:AH1")
        ws.merge_cells("A2:AH2")
        
        ws.merge_cells("A4:A5") # STT
        ws.merge_cells("B4:B5") # MA
        ws.merge_cells("C4:C5") # TAn
        ws.merge_cells("D4:D5") # So dien thoai
        ws.merge_cells("E4:E5") # Dia chi
        ws.merge_cells("F4:F5") # Ten Zalo
        
        ws.merge_cells("G4:L4") # Hoa don
        ws.merge_cells("M4:Q4") # Danh gia
        ws.merge_cells("R4:T4") # Chi phi
        ws.merge_cells("U4:X4") # Tuyen xe
        ws.merge_cells("Y4:AF4") # Nang luc
        ws.merge_cells("AG4:AG5") # Muc gia
        ws.merge_cells("AH4:AH5") # Nguoi quan ly
        
        title_font = Font(size=16, bold=True, color="FF0000")
        subtitle_font = Font(size=14, bold=True, color="0000FF")
        align_center = Alignment(horizontal="center", vertical="center")
        
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = align_center
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row_num in [4, 5]:
            for col_num in range(1, len(headers_l2) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
                
        row_idx = 6
        stt = 1
        
        fill_even = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        fill_odd = PatternFill(start_color='E8F0F8', end_color='E8F0F8', fill_type='solid') # Lighter blue/gray
        
        for p_idx, p in enumerate(partners):
            routes = p.bao_gia.filter(is_deleted=False).order_by('id')
            p_start_row = row_idx
            p_co_hop_dong = "Có" if p.co_hop_dong else "Không"
            current_fill = fill_even if p_idx % 2 == 0 else fill_odd
            
            if not routes.exists():
                row_data = [
                    stt, p.ma_nha_xe, p.ten_nha_xe, getattr(p, 'so_dien_thoai', ''), p.dia_chi, getattr(p, 'ten_zalo', ''),
                    p.ten_don_vi_xuat_hoa_don, p.dia_chi_xuat_hoa_don, p.ma_so_thue, p.so_tai_khoan, p.nguoi_dai_dien, p.ngan_hang,
                    p.thoi_han_thanh_toan, p.tinh_san_sang, p_co_hop_dong, p.chat_luong, p.ghi_chu,
                    p.phi_rot_diem, p.phi_boc_xep, p.phi_di_qua_tai,
                    "", "", "", "",
                    "", "", "", "", "", "", "", "",
                    "", getattr(p, 'nguoi_quan_ly', '')
                ]
                ws.append(row_data)
                for col_num in range(1, len(headers_l2) + 1):
                    c = ws.cell(row=row_idx, column=col_num)
                    c.border = thin_border
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                    c.fill = current_fill
                    if col_num in [18, 19, 20, 33]: 
                        c.number_format = '#,##0'
                        c.alignment = Alignment(vertical="center", horizontal="right")
                row_idx += 1
            else:
                for idx, r in enumerate(routes):
                    chieu_di = f"{r.di_1_hay_2_chieu} chiều" if r.di_1_hay_2_chieu else ""
                    nhieu_diem = "Có" if r.co_di_nhieu_diem_khong else "Không"
                    ghep_hang = "Có" if r.co_ghep_hang_khong else "Không"
                    qua_tai = "Có" if r.co_chiu_qua_tai_khong else "Không"
                    
                    row_data = [
                        stt, p.ma_nha_xe, p.ten_nha_xe, getattr(p, 'so_dien_thoai', ''), p.dia_chi, getattr(p, 'ten_zalo', ''),
                        p.ten_don_vi_xuat_hoa_don, p.dia_chi_xuat_hoa_don, p.ma_so_thue, p.so_tai_khoan, p.nguoi_dai_dien, p.ngan_hang,
                        p.thoi_han_thanh_toan, p.tinh_san_sang, p_co_hop_dong, p.chat_luong, p.ghi_chu,
                        p.phi_rot_diem, p.phi_boc_xep, p.phi_di_qua_tai,
                        r.tuyen.tinh_nhan, r.tuyen.huyen_nhan, r.tuyen.tinh_giao, r.tuyen.huyen_giao,
                        chieu_di, nhieu_diem, ghep_hang, qua_tai, 
                        r.tai_trong_tan, r.so_khoi, r.kich_thuoc, r.loai_thung,
                        r.muc_gia_chap_nhan, getattr(p, 'nguoi_quan_ly', '')
                    ]
                    ws.append(row_data)
                    for col_num in range(1, len(headers_l2) + 1):
                        c = ws.cell(row=row_idx, column=col_num)
                        c.border = thin_border
                        c.alignment = Alignment(vertical="center", wrap_text=True)
                        c.fill = current_fill
                        if col_num in [18, 19, 20, 33]: 
                            c.number_format = '#,##0'
                            c.alignment = Alignment(vertical="center", horizontal="right")
                    row_idx += 1
                    
            if row_idx - 1 > p_start_row:
                for c_idx in range(1, 21):
                    col_letter = openpyxl.utils.get_column_letter(c_idx)
                    ws.merge_cells(f"{col_letter}{p_start_row}:{col_letter}{row_idx - 1}")
                    # Re-apply alignment for merged cells
                    horz = "left"
                    if c_idx == 1: horz = "right"
                    elif c_idx in [18, 19, 20]: horz = "right"
                    ws.cell(row=p_start_row, column=c_idx).alignment = Alignment(horizontal=horz, vertical="center", wrap_text=True)
                    
            stt += 1

        autofit_columns(ws, skip_rows=2)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if partner_id:
            response['Content-Disposition'] = f'attachment; filename="Doi_tac_{p.ma_nha_xe or p.id}.xlsx"'
        else:
            response['Content-Disposition'] = 'attachment; filename="Danh_sach_doi_tac_thue_ngoai.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        from django.http import HttpResponse
        return HttpResponse(f"Lỗi: {str(e)}", status=500)

def api_export_partner_template(request):
    try:
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from core.constants import PROVINCES_DATA
        import re
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DanhSachDoiTac"
        
        # Create hidden sheet
        hidden_ws = wb.create_sheet("ListData")
        hidden_ws.sheet_state = 'hidden'
        
        # Setup basic list validations
        san_sang_list = ['Cao', 'Trung bình', 'Thấp']
        co_khong_list = ['Có', 'Không']
        loai_thung_list = ['Thùng kín', 'Thùng bạt', 'Khác']
        so_chieu_di_list = ['1 chiều', '2 chiều']
        
        from core.models import CauHinhLoaiXe
        active_types = ["1.25T", "2.5T", "3.5T", "5T", "7T", "8T", "9T", "15T"]
        tai_trong_list = active_types + ['LTL']
        
        tai_trong_mapping = {}
        for lx in active_types:
            ch = CauHinhLoaiXe.objects.filter(loai_xe=lx).first()
            if ch:
                tai_trong_mapping[lx] = ch.get_so_khoi()
            else:
                default_max = {"1.25T": 9, "2.5T": 13, "3.5T": 18, "5T": 23, "7T": 35, "8T": 54, "9T": 40, "15T": 54}
                tai_trong_mapping[lx] = str(default_max[lx])
        tai_trong_mapping['LTL'] = 'LTL'
        so_khoi_list = list(tai_trong_mapping.values())
        col_idx = 1
        for tt, sk in tai_trong_mapping.items():
            tt_clean = "_" + tt.replace('.', '_')
            hidden_ws.cell(row=1, column=col_idx, value=tt_clean)
            hidden_ws.cell(row=2, column=col_idx, value=sk)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ref = f"ListData!${col_letter}$2:${col_letter}$2"
            wb.create_named_range(tt_clean, None, ref)
            col_idx += 1
            
        lists = [
            ("SanSang", san_sang_list),
            ("CoKhong", co_khong_list),
            ("LoaiThung", loai_thung_list),
            ("SoChieuDi", so_chieu_di_list),
            ("TaiTrong", tai_trong_list),
            ("SoKhoi", so_khoi_list)
        ]
        
        for name, items in lists:
            hidden_ws.cell(row=1, column=col_idx, value=name)
            for i, v in enumerate(items, 2):
                hidden_ws.cell(row=i, column=col_idx, value=v)
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ref = f"ListData!${col_letter}$2:${col_letter}${len(items)+1}"
            wb.create_named_range(name, None, ref)
            col_idx += 1
            
        def clean_name(s):
            return s.replace(" ", "").replace("-", "").replace(".", "")

        # Provinces and Districts
        provinces = [p['name'].replace("Tỉnh ", "").replace("Thành phố ", "") for p in PROVINCES_DATA]
        hidden_ws.cell(row=1, column=col_idx, value="Provinces")
        for i, p in enumerate(provinces, 2):
            hidden_ws.cell(row=i, column=col_idx, value=p)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ref = f"ListData!${col_letter}$2:${col_letter}${len(provinces)+1}"
        wb.create_named_range("Provinces", None, ref)
        col_idx += 1
        
        for p_data in PROVINCES_DATA:
            p_name = p_data['name'].replace("Tỉnh ", "").replace("Thành phố ", "")
            districts = [d['name'] for d in p_data['districts']]
            p_name_clean = clean_name(p_name)
            if not p_name_clean: p_name_clean = f"Prov{col_idx}"
            hidden_ws.cell(row=1, column=col_idx, value=p_name_clean)
            for i, d in enumerate(districts, 2):
                hidden_ws.cell(row=i, column=col_idx, value=d)
                
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ref = f"ListData!${col_letter}$2:${col_letter}${len(districts)+1}"
            wb.create_named_range(p_name_clean, None, ref)
            col_idx += 1
        
        ws.append(["CÔNG TY CỔ PHẦN NHẤT PHONG VẬN"])
        ws.append(["DANH SÁCH ĐỐI TÁC THUÊ NGOÀI VÀ BẢNG GIÁ"])
        ws.append(["Lưu ý: Bắt buộc nhập Mã nhà xe. Tỉnh nhận & Tỉnh giao bắt buộc nếu có Tuyến xe. Các dòng có cùng Mã nhà xe sẽ được gộp vào chung 1 đối tác."])
        
        # Headers level 1
        headers_l1 = [
            "STT", "Mã nhà xe", "Tên nhà xe", "Số điện thoại", "Địa chỉ", "Tên Zalo",
            "THÔNG TIN XUẤT HOÁ ĐƠN", "", "", "", "", "",
            "ĐÁNH GIÁ & THANH TOÁN", "", "", "", "",
            "CÁC LOẠI CHI PHÍ NHÀ XE", "", "",
            "TUYẾN XE", "", "", "",
            "THÔNG TIN XE & NĂNG LỰC", "", "", "", "", "", "", "",
            "MỨC GIÁ CHẤP NHẬN", "NGƯỜI QUẢN LÝ"
        ]
        
        # Headers level 2
        headers_l2 = [
            "STT", "Mã nhà xe", "Tên nhà xe", "Số điện thoại", "Địa chỉ", "Tên Zalo",
            "Tên đơn vị", "Địa chỉ HĐ", "Mã số thuế", "Số tài khoản", "Người đại diện", "Ngân hàng",
            "Thời hạn thanh toán", "Tính sẵn sàng", "Có hợp đồng không", "Chất lượng", "Ghi chú",
            "Phí rớt điểm", "Phí bốc xếp", "Phí đi quá tải",
            "Tỉnh nhận", "Huyện nhận", "Tỉnh giao", "Huyện giao",
            "Số chiều đi", "Đi nhiều điểm", "Ghép hàng", "Chịu quá tải", "Tải trọng (Tấn)", "Số khối", "Kích thước", "Loại thùng",
            "Mức giá (VNĐ)", "Người quản lý"
        ]
        
        ws.append(headers_l1)
        ws.append(headers_l2)
        
        ws.merge_cells("A1:AH1")
        ws.merge_cells("A2:AH2")
        ws.merge_cells("A3:AH3")
        
        ws.merge_cells("A4:A5") # STT
        ws.merge_cells("B4:B5") # MA
        ws.merge_cells("C4:C5") # TAn
        ws.merge_cells("D4:D5") # So dien thoai
        ws.merge_cells("E4:E5") # Dia chi
        ws.merge_cells("F4:F5") # Ten Zalo
        
        ws.merge_cells("G4:L4") # Hoa don
        ws.merge_cells("M4:Q4") # Danh gia
        ws.merge_cells("R4:T4") # Chi phi
        ws.merge_cells("U4:X4") # Tuyen xe
        ws.merge_cells("Y4:AF4") # Nang luc
        ws.merge_cells("AG4:AG5") # Muc gia
        ws.merge_cells("AH4:AH5") # Nguoi quan ly
        
        title_font = Font(size=16, bold=True, color="FF0000")
        subtitle_font = Font(size=14, bold=True, color="0000FF")
        note_font = Font(size=11, italic=True, color="FF0000")
        align_center = Alignment(horizontal="center", vertical="center")
        
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = align_center
        ws["A3"].font = note_font
        ws["A3"].alignment = align_center
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row_num in [4, 5]:
            for col_num in range(1, len(headers_l2) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
                
        # Applying Data Validations
        dv_sansang = DataValidation(type="list", formula1="=SanSang", allow_blank=True)
        dv_cokhong = DataValidation(type="list", formula1="=CoKhong", allow_blank=True)
        dv_loaithung = DataValidation(type="list", formula1="=LoaiThung", allow_blank=True)
        dv_sochieudi = DataValidation(type="list", formula1="=SoChieuDi", allow_blank=True)
        dv_taitrong = DataValidation(type="list", formula1="=TaiTrong", allow_blank=True)
        dv_sokhoi = DataValidation(type="list", formula1='=INDIRECT("_" & SUBSTITUTE(AC6, ".", "_"))', allow_blank=True)
        dv_province = DataValidation(type="list", formula1="=Provinces", allow_blank=True)
        
        # Note: INDIRECT needs cell reference without absolute, but openpyxl applies it correctly to a range
        dv_district_nhan = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(U6, " ", ""), "-", ""), ".", ""))', allow_blank=True)
        dv_district_giao = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(W6, " ", ""), "-", ""), ".", ""))', allow_blank=True)
        
        ws.add_data_validation(dv_sansang)
        ws.add_data_validation(dv_cokhong)
        ws.add_data_validation(dv_loaithung)
        ws.add_data_validation(dv_sochieudi)
        ws.add_data_validation(dv_taitrong)
        ws.add_data_validation(dv_sokhoi)
        ws.add_data_validation(dv_province)
        ws.add_data_validation(dv_district_nhan)
        ws.add_data_validation(dv_district_giao)
        
        dv_sansang.add("N6:N1000")
        dv_cokhong.add("O6:O1000")
        dv_cokhong.add("Z6:Z1000")
        dv_cokhong.add("AA6:AA1000")
        dv_cokhong.add("AB6:AB1000")
        dv_sochieudi.add("Y6:Y1000")
        dv_taitrong.add("AC6:AC1000")
        dv_sokhoi.add("AD6:AD1000")
        dv_loaithung.add("AF6:AF1000")
        dv_province.add("U6:U1000")
        dv_province.add("W6:W1000")
        dv_district_nhan.add("V6:V1000")
        dv_district_giao.add("X6:X1000")

        widths = [5, 15, 25, 15, 25, 15, 25, 30, 15, 15, 15, 15, 15, 12, 12, 10, 20, 15, 15, 15, 15, 15, 15, 15, 12, 12, 12, 12, 15, 10, 20, 15, 15, 20]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            
        # Format empty rows for prices (columns 18, 19, 20, 33)
        for row in range(6, 1001):
            for col in [18, 19, 20, 33]:
                c = ws.cell(row=row, column=col)
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right', vertical='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Danh_sach_doi_tac_thue_ngoai.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        from django.http import HttpResponse
        return HttpResponse(f"Lỗi: {str(e)}", status=500)

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

@csrf_exempt
def api_import_partners_excel(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'Không tìm thấy file upload.'})
            
        file = request.FILES['file']
        import openpyxl
        from core.models import DoiTac, TuyenXe, BaoGiaThongTinXe
        from django.utils import timezone
        
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        count_partner = 0
        count_route = 0
        partner_map = {}
        
        # New Layout starting at row 6
        for row_idx in range(6, ws.max_row + 1):
            row = [cell.value for cell in ws[row_idx]]
            if not any(row): continue
            
            # Map columns according to new headers_l2
            # 1: Mã, 2: Tên, 3: Số ĐT, 4: Địa chỉ, 5: Tên Zalo
            # 6: Tên đv, 7: Địa chỉ HĐ, 8: MST, 9: Số TK, 10: Ng đại diện, 11: Ngân hàng
            # 12: Thời hạn TT, 13: Tính sẵn sàng, 14: Có HĐ, 15: Chất lượng, 16: Ghi chú
            # 17: Phí rớt, 18: Phí bốc xếp, 19: Phí quá tải
            # 20: Tỉnh nhận, 21: Huyện nhận, 22: Tỉnh giao, 23: Huyện giao
            # 24: Số chiều đi, 25: Nhiều điểm, 26: Ghép hàng, 27: Chịu quá tải
            # 28: Tải trọng, 29: Số khối, 30: Kích thước, 31: Loại thùng, 32: Mức giá
            # 33: Người quản lý
            
            ma_nha_xe_input = str(row[1]).strip() if len(row) > 1 and row[1] else None
            if ma_nha_xe_input and ma_nha_xe_input.lower() == 'none': ma_nha_xe_input = None
            
            so_dien_thoai_input = str(row[3]).strip() if len(row) > 3 and row[3] else None
            nguoi_quan_ly_input = str(row[33]).strip() if len(row) > 33 and row[33] else None
            
            ma_so_thue_input = str(row[8]).strip() if len(row) > 8 and row[8] else None
            mst_clean = ""
            if ma_so_thue_input:
                mst_clean = ma_so_thue_input.replace('-', '').replace(' ', '')
                if not mst_clean.isdigit() or len(mst_clean) not in [10, 13]:
                    return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {row_idx}: Mã số thuế "{ma_so_thue_input}" không hợp lệ (phải có 10 hoặc 13 số).'})
            
            # Identify by MST ONLY globally
            p = None
            if mst_clean:
                p_global = DoiTac.objects.filter(ma_so_thue=mst_clean).first()
                if p_global:
                    if not request.user.is_superuser and p_global.nguoi_quan_ly != request.user.username:
                        manager_name = p_global.nguoi_quan_ly if p_global.nguoi_quan_ly else "Admin"
                        return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {row_idx}: Nhà xe với Mã số thuế {mst_clean} đã tồn tại do {manager_name} quản lý.'})
                    else:
                        p = p_global
                        
            if not p:
                p = DoiTac()
                if not request.user.is_superuser:
                    p.nguoi_quan_ly = nguoi_quan_ly_input if nguoi_quan_ly_input else request.user.username
                else:
                    if nguoi_quan_ly_input: p.nguoi_quan_ly = nguoi_quan_ly_input
                count_partner += 1
                
            # Assign ma_nha_xe directly
            if ma_nha_xe_input:
                p.ma_nha_xe = ma_nha_xe_input
            
            # We don't skip if they are updating existing partner map in the same excel
            partner_key = p.pk if p.pk else f"new_{count_partner}"
            
            if partner_key not in partner_map:
                if len(row) > 2 and row[2]: p.ten_nha_xe = str(row[2]).strip()
                if request.user.is_superuser and nguoi_quan_ly_input: p.nguoi_quan_ly = nguoi_quan_ly_input
                if so_dien_thoai_input: p.so_dien_thoai = so_dien_thoai_input
                if len(row) > 4 and row[4]: p.dia_chi = str(row[4]).strip()
                if len(row) > 5 and row[5]: p.ten_zalo = str(row[5]).strip()
                if len(row) > 6 and row[6]: p.ten_don_vi_xuat_hoa_don = str(row[6]).strip()
                if len(row) > 7 and row[7]: p.dia_chi_xuat_hoa_don = str(row[7]).strip()
                if mst_clean: p.ma_so_thue = mst_clean
                if len(row) > 9 and row[9]: p.so_tai_khoan = str(row[9]).strip()
                if len(row) > 10 and row[10]: p.nguoi_dai_dien = str(row[10]).strip()
                if len(row) > 11 and row[11]: p.ngan_hang = str(row[11]).strip()
                
                if len(row) > 12 and row[12]: p.thoi_han_thanh_toan = str(row[12]).strip()
                if len(row) > 13 and row[13]: p.tinh_san_sang = str(row[13]).strip()
                
                co_hd = str(row[14]).strip().lower() if len(row) > 14 and row[14] else ""
                p.co_hop_dong = "cA3" in co_hd or "co" in co_hd
                
                if len(row) > 15 and row[15]:
                    try: p.chat_luong = int(row[15])
                    except: return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {row_idx}: Chất lượng "{row[15]}" không hợp lệ (phải là số).'})
                
                if len(row) > 16 and row[16]: p.ghi_chu = str(row[16]).strip()
                
                if len(row) > 17 and row[17]:
                    try: p.phi_rot_diem = float(row[17])
                    except: return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {row_idx}: Phí rớt điểm "{row[17]}" không hợp lệ (phải là số).'})
                if len(row) > 18 and row[18]:
                    try: p.phi_boc_xep = float(row[18])
                    except: return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {row_idx}: Phí bốc xếp "{row[18]}" không hợp lệ (phải là số).'})
                if len(row) > 19 and row[19]:
                    try: p.phi_di_qua_tai = float(row[19])
                    except: return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {row_idx}: Phí đi quá tải "{row[19]}" không hợp lệ (phải là số).'})
                
                p.save()
            
            # Now add route
            if len(row) > 20 and any(row[20:33]):
                tinh_nhan = str(row[20]).strip() if len(row) > 20 and row[20] else None
                tinh_giao = str(row[22]).strip() if len(row) > 22 and row[22] else None
                
                if tinh_nhan and tinh_giao and str(tinh_nhan).lower() != 'none' and str(tinh_giao).lower() != 'none':
                    huyen_nhan = str(row[21]).strip() if len(row) > 21 and row[21] else ""
                    huyen_giao = str(row[23]).strip() if len(row) > 23 and row[23] else ""
                    if huyen_nhan.lower() == 'none': huyen_nhan = ""
                    if huyen_giao.lower() == 'none': huyen_giao = ""
                    
                    tuyen = TuyenXe.objects.filter(
                        tinh_nhan__iexact=tinh_nhan, huyen_nhan__iexact=huyen_nhan,
                        tinh_giao__iexact=tinh_giao, huyen_giao__iexact=huyen_giao
                    ).first()
                    
                    if not tuyen:
                        from django.utils.text import slugify
                        ma = f"{slugify(tinh_nhan)}-{slugify(tinh_giao)}"[:40]
                        base_ma = ma
                        counter = 1
                        while TuyenXe.objects.filter(ma_tuyen=ma).exists():
                            ma = f"{base_ma}-{counter}"
                            counter += 1
                        tuyen = TuyenXe.objects.create(
                            ma_tuyen=ma, tinh_nhan=tinh_nhan, huyen_nhan=huyen_nhan,
                            tinh_giao=tinh_giao, huyen_giao=huyen_giao
                        )
                    
                    bg = BaoGiaThongTinXe.objects.filter(doi_tac=p, tuyen=tuyen).first()
                    if not bg:
                        bg = BaoGiaThongTinXe(doi_tac=p, tuyen=tuyen)
                        
                    chieu_di = str(row[24]).strip().lower() if len(row) > 24 and row[24] else ""
                    bg.di_1_hay_2_chieu = 2 if "2" in chieu_di else 1
                    
                    n_diem = str(row[25]).strip().lower() if len(row) > 25 and row[25] else ""
                    bg.co_di_nhieu_diem_khong = "có" in n_diem or "co" in n_diem
                    
                    g_hang = str(row[26]).strip().lower() if len(row) > 26 and row[26] else ""
                    bg.co_ghep_hang_khong = "có" in g_hang or "co" in g_hang
                    
                    q_tai = str(row[27]).strip().lower() if len(row) > 27 and row[27] else ""
                    bg.co_chiu_qua_tai_khong = "có" in q_tai or "co" in q_tai
                    
                    tt = str(row[28]).strip() if len(row) > 28 and row[28] else ""
                    import re
                    nums = re.findall(r"[-+]?\d*\.\d+|\d+", tt)
                    if nums:
                        bg.tai_trong_tan = float(nums[0])
                    else:
                        bg.tai_trong_tan = 0.0
                    
                    bg.so_khoi = str(row[29]).strip() if len(row) > 29 and row[29] and str(row[29]).lower() != 'none' else ""
                    bg.kich_thuoc = str(row[30]).strip() if len(row) > 30 and row[30] and str(row[30]).lower() != 'none' else ""
                    bg.loai_thung = str(row[31]).strip() if len(row) > 31 and row[31] and str(row[31]).lower() != 'none' else ""
                    
                    if len(row) > 32 and row[32]:
                        try: 
                            bg.muc_gia_chap_nhan = float(row[32])
                        except: 
                            return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {row_idx}: Mức giá "{row[32]}" không hợp lệ (phải là số).'})
                    else:
                        bg.muc_gia_chap_nhan = 0.0
                    
                    bg.save()
                    count_route += 1
                    
            partner_map[partner_key] = p

        return JsonResponse({'success': True, 'message': f'Nhập thành công. Cập nhật {len(partner_map)} nhà xe, tạo mới {count_partner} nhà xe, lưu {count_route} tuyến.'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_search_prices(request):
    try:
        tinh_nhan = request.GET.get('tinh_nhan', '').strip()
        huyen_nhan = request.GET.get('huyen_nhan', '').strip()
        tinh_giao = request.GET.get('tinh_giao', '').strip()
        huyen_giao = request.GET.get('huyen_giao', '').strip()
        loai_xe = request.GET.get('loai_xe', '').strip()
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
        
        
        # Build base quote query
        if request.user.is_superuser:
            bg_qs = BaoGiaThongTinXe.objects.select_related('doi_tac', 'tuyen').all()
        else:
            bg_qs = BaoGiaThongTinXe.objects.select_related('doi_tac', 'tuyen').filter(doi_tac__nguoi_quan_ly=request.user.username)
        
        if tinh_nhan:
            bg_qs = bg_qs.filter(tuyen__tinh_nhan=tinh_nhan)
        if huyen_nhan:
            huyen_nhan = clean_huyen(huyen_nhan)
            bg_qs = bg_qs.filter(tuyen__huyen_nhan=huyen_nhan)
        if tinh_giao:
            bg_qs = bg_qs.filter(tuyen__tinh_giao=tinh_giao)
        if huyen_giao:
            huyen_giao = clean_huyen(huyen_giao)
            bg_qs = bg_qs.filter(tuyen__huyen_giao=huyen_giao)
        if loai_xe:
            try:
                loai_xe_val = float(loai_xe.lower().replace('t', '').strip())
                bg_qs = bg_qs.filter(tai_trong_tan=loai_xe_val)
            except ValueError:
                pass
                
        so_khoi = request.GET.get('so_khoi', '').strip()
        if so_khoi:
            bg_qs = bg_qs.filter(so_khoi=so_khoi)
                
        bg_qs = bg_qs.order_by('muc_gia_chap_nhan')
        
        # Pre-fetch Base Prices for quick lookup
        gcs_dict = {}
        for g in GiaCoSo.objects.select_related('tuyen').all():
            key = (g.tuyen.id, g.loai_xe)
            gcs_dict[key] = g.gia_co_so

        def get_lx_str(tt):
            if tt.is_integer():
                return f"{int(tt)}T"
            return f"{tt}T"

        results = []
        is_exact_search = bool(tinh_nhan and tinh_giao and (loai_xe or so_khoi))
        exact_base_price = None

        for bg in bg_qs:
            lx_str = get_lx_str(bg.tai_trong_tan)
            key = (bg.tuyen.id, lx_str)
            g_val = gcs_dict.get(key)
            
            # The user requested to show all results regardless of base price
            # if g_val is not None:
            #     if bg.muc_gia_chap_nhan > g_val:
            #         continue
                
            # For exact search info box
            if is_exact_search and g_val is not None and not exact_base_price:
                # We just find the first match to display if they did an exact search
                g_obj = GiaCoSo.objects.filter(tuyen=bg.tuyen, loai_xe=lx_str).first()
                if g_obj:
                    exact_base_price = {
                        'gia_co_so': g_obj.gia_co_so,
                        'so_khoi': g_obj.so_khoi or '-',
                        'ngay_ap_dung': g_obj.ngay_ap_dung.strftime('%d/%m/%Y') if g_obj.ngay_ap_dung else '-'
                    }

            results.append({
                'bg_id': bg.id,
                'doi_tac_id': bg.doi_tac.id,
                'doi_tac_ten': bg.doi_tac.ten_nha_xe,
                'doi_tac_sdt': bg.doi_tac.so_dien_thoai or 'Không có',
                'tuyen': f"{bg.tuyen.huyen_nhan + ', ' if bg.tuyen.huyen_nhan else ''}{bg.tuyen.tinh_nhan} ➔ {bg.tuyen.huyen_giao + ', ' if bg.tuyen.huyen_giao else ''}{bg.tuyen.tinh_giao}",
                'loai_xe': lx_str,
                'so_khoi': bg.so_khoi or '-',
                'kich_thuoc': bg.kich_thuoc or '-',
                'loai_thung': bg.loai_thung or '-',
                'muc_gia': bg.muc_gia_chap_nhan,
                'ghep_hang': 'Có' if bg.co_ghep_hang_khong else 'Không',
                'chieu_di': f"{bg.di_1_hay_2_chieu} chiều",
                'gia_co_so': g_val,
            })
            
        return JsonResponse({'success': True, 'is_exact_search': is_exact_search, 'base_price': exact_base_price, 'results': results})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def proposal_list_view(request):
    if request.user.is_superuser:
        proposals = DeXuatGiaCoSo.objects.all().order_by('-ngay_tao')
    else:
        proposals = DeXuatGiaCoSo.objects.filter(nguoi_de_xuat=request.user.username).order_by('-ngay_tao')
        DeXuatGiaCoSo.objects.filter(nguoi_de_xuat=request.user.username, is_read=False).exclude(trang_thai='ChoDuyet').update(is_read=True)
    
    tinh_nhan_qs = TuyenXe.objects.values_list('tinh_nhan', flat=True).distinct()
    tinh_giao_qs = TuyenXe.objects.values_list('tinh_giao', flat=True).distinct()
    tinh_thanh = sorted(list(set(tinh_nhan_qs) | set(tinh_giao_qs)))
    loai_xes = sorted(list(GiaCoSo.objects.values_list('loai_xe', flat=True).distinct()))
    
    return render(request, 'core/proposal_list.html', {
        'proposals': proposals,
        'tinh_thanh': tinh_thanh,
        'loai_xes': loai_xes
    })

@login_required
def api_create_proposal(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        data = json.loads(request.body)
        tinh_nhan = data.get('tinh_nhan', '').strip()
        huyen_nhan = data.get('huyen_nhan', '').strip()
        tinh_giao = data.get('tinh_giao', '').strip()
        huyen_giao = data.get('huyen_giao', '').strip()
        loai_xe = data.get('loai_xe', '').strip()
        so_khoi = data.get('so_khoi', '').strip()
        gia_de_xuat = data.get('gia_de_xuat')
        ly_do = data.get('ly_do')
        
        prefixes_tinh = ["Tỉnh ", "Thành phố "]
        for p in prefixes_tinh:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        huyen_nhan = clean_huyen(huyen_nhan)
        huyen_giao = clean_huyen(huyen_giao)
        
        if not (tinh_nhan and tinh_giao and loai_xe and gia_de_xuat and ly_do):
            return JsonResponse({'success': False, 'error': 'Thiếu thông tin bắt buộc'})
            
        tuyen, created = TuyenXe.objects.get_or_create(
            tinh_nhan=tinh_nhan, huyen_nhan=huyen_nhan, tinh_giao=tinh_giao, huyen_giao=huyen_giao,
            defaults={'ma_tuyen': f"T-DX-{int(time.time())}"}
        )
        
        gcs = GiaCoSo.objects.filter(tuyen=tuyen, loai_xe=loai_xe).first()
        gia_hien_tai = gcs.gia_co_so if gcs else None
        
        gia_de_xuat_clean = str(gia_de_xuat).replace(',', '').replace(' ', '')
        
        DeXuatGiaCoSo.objects.create(
            tuyen=tuyen, loai_xe=loai_xe, so_khoi=so_khoi,
            gia_hien_tai=gia_hien_tai, gia_de_xuat=parse_vn_number(gia_de_xuat_clean),
            ly_do_de_xuat=ly_do, nguoi_de_xuat=request.user.username
        )
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_approve_proposal(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
        
    try:
        proposal = DeXuatGiaCoSo.objects.get(pk=pk)
        if proposal.trang_thai != 'ChoDuyet':
            return JsonResponse({'success': False, 'error': 'Đề xuất này đã được xử lý'})
            
        proposal.trang_thai = 'DaDuyet'
        proposal.is_read = False
        proposal.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_reject_proposal(request, pk):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Unauthorized'})
        
    try:
        data = json.loads(request.body)
        proposal = DeXuatGiaCoSo.objects.get(pk=pk)
        if proposal.trang_thai != 'ChoDuyet':
            return JsonResponse({'success': False, 'error': 'Đề xuất này đã được xử lý'})
            
        proposal.trang_thai = 'TuChoi'
        proposal.save()
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_get_base_price(request):
    try:
        tinh_nhan = request.GET.get('tinh_nhan', '').strip()
        tinh_giao = request.GET.get('tinh_giao', '').strip()
        loai_xe = request.GET.get('loai_xe', '').strip()
        
        prefixes = ["Tỉnh ", "Thành phố "]
        for p in prefixes:
            if tinh_nhan.startswith(p): tinh_nhan = tinh_nhan[len(p):]
            if tinh_giao.startswith(p): tinh_giao = tinh_giao[len(p):]
            
        gcs = GiaCoSo.objects.filter(tuyen__tinh_nhan=tinh_nhan, tuyen__tinh_giao=tinh_giao, loai_xe=loai_xe).first()
        if gcs:
            return JsonResponse({'success': True, 'gia_co_so': gcs.gia_co_so, 'so_khoi': gcs.so_khoi})
        return JsonResponse({'success': True, 'gia_co_so': None})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_delete_partner(request, pk):
    if request.method != 'POST': return JsonResponse({'success': False})
    try:
        if request.user.is_superuser: p = DoiTac.objects.get(pk=pk)
        else: p = DoiTac.objects.get(pk=pk, nguoi_quan_ly=request.user.username)
        p.is_deleted = True; p.save()
        return JsonResponse({'success': True})
    except Exception as e: return JsonResponse({'success': False})

@login_required
def api_bulk_delete_partners(request):
    if request.method != 'POST': return JsonResponse({'success': False})
    try:
        ids_str = request.POST.get('ids', '')
        if not ids_str: return JsonResponse({'success': False, 'error': 'No ids provided'})
        ids = [int(i.strip()) for i in ids_str.split(',') if i.strip().isdigit()]
        
        if request.user.is_superuser:
            DoiTac.objects.filter(id__in=ids).update(is_deleted=True)
        else:
            DoiTac.objects.filter(id__in=ids, nguoi_quan_ly=request.user.username).update(is_deleted=True)
            
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_delete_base_price(request, pk):
    if request.method != 'POST': return JsonResponse({'success': False})
    try:
        if not request.user.is_superuser: return JsonResponse({'success': False})
        p = GiaCoSo.objects.get(pk=pk)
        p.is_deleted = True; p.save()
        return JsonResponse({'success': True})
    except Exception as e: return JsonResponse({'success': False})

@login_required
def api_check_duplicate_partner_routes(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})
        
    try:
        data = json.loads(request.body)
        partner_id = data.get('partner_id')
        new_routes = data.get('new_routes', [])
        
        if not partner_id:
            return JsonResponse({'success': True, 'has_duplicate': False})
            
        partner = get_object_or_404(DoiTac, pk=partner_id)
        
        for route in new_routes:
            tinh_n = route.get('tinh_nhan', '').replace("Tỉnh ", "").replace("Thành phố ", "").strip()
            huyen_n = clean_huyen(route.get('huyen_nhan', ''))
            tinh_g = route.get('tinh_giao', '').replace("Tỉnh ", "").replace("Thành phố ", "").strip()
            huyen_g = clean_huyen(route.get('huyen_giao', ''))
            
            try:
                tai_trong = float(route.get('tai_trong_tan', '0'))
            except ValueError:
                tai_trong = 0.0
            so_khoi = route.get('so_khoi', '')
            if not so_khoi: so_khoi = None
            
            loai_thung = route.get('loai_thung', '')
            ghep_hang = True if route.get('co_ghep_hang_khong') == '1' else False
            qua_tai = True if route.get('co_chiu_qua_tai_khong') == '1' else False
            nhieu_diem = True if route.get('co_di_nhieu_diem_khong') == '1' else False
            
            chieu_di_str = str(route.get('di_1_hay_2_chieu', '1'))
            chieu_di = int(chieu_di_str) if chieu_di_str.isdigit() else 1
            
            # Find duplicate
            existing = BaoGiaThongTinXe.objects.filter(
                doi_tac=partner,
                tuyen__tinh_nhan=tinh_n,
                tuyen__huyen_nhan=huyen_n,
                tuyen__tinh_giao=tinh_g,
                tuyen__huyen_giao=huyen_g,
                tai_trong_tan=tai_trong,
                loai_thung=loai_thung,
                co_ghep_hang_khong=ghep_hang,
                co_chiu_qua_tai_khong=qua_tai,
                co_di_nhieu_diem_khong=nhieu_diem,
                di_1_hay_2_chieu=chieu_di
            )
            
            kich_thuoc = route.get('kich_thuoc', '')
            if kich_thuoc:
                existing = existing.filter(kich_thuoc=kich_thuoc)
            else:
                existing = existing.filter(kich_thuoc='')
                
            if so_khoi is not None:
                existing = existing.filter(so_khoi=so_khoi)
            else:
                existing = existing.filter(so_khoi__isnull=True)
                
            if existing.exists():
                duplicate = existing.first()
                return JsonResponse({
                    'success': True,
                    'has_duplicate': True,
                    'duplicate_route_id': duplicate.id,
                    'error': f'Tuyến {tinh_n} đi {tinh_g} (Tải trọng: {tai_trong} tấn) đã tồn tại trong báo giá của đối tác này.'
                })
                
        return JsonResponse({'success': True, 'has_duplicate': False})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_bulk_update_base_prices(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Chỉ admin mới có quyền chỉnh sửa giá cơ sở'})
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        data = json.loads(request.body)
        ids = data.get('ids', [])
        percentage = data.get('percentage')
        ngay_ap_dung_str = data.get('ngay_ap_dung')
        ly_do_doi = data.get('ly_do_doi', 'Cập nhật giá hàng loạt')
        
        if percentage is None or percentage == '':
            return JsonResponse({'success': False, 'error': 'Thiếu tỷ lệ phần trăm'})
            
        percentage = float(percentage)
        
        ngay_ap_dung = None
        if ngay_ap_dung_str:
            from datetime import datetime
            ngay_ap_dung = datetime.strptime(ngay_ap_dung_str, '%Y-%m-%d').date()
            
        for pk in ids:
            try:
                g = GiaCoSo.objects.get(pk=pk)
                old_price = g.gia_co_so
                new_price = old_price * (1 + percentage / 100)
                g.gia_co_so = new_price
                if ngay_ap_dung:
                    g.ngay_ap_dung = ngay_ap_dung
                g._ly_do_doi = ly_do_doi
                g.save()
            except GiaCoSo.DoesNotExist:
                continue
                
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_check_new_proposals(request):
    if request.user.is_superuser:
        count = DeXuatGiaCoSo.objects.filter(trang_thai='ChoDuyet').count()
        return JsonResponse({'success': True, 'count': count, 'role': 'admin'})
    else:
        # Get count of unread approved/rejected proposals
        unread_proposals = DeXuatGiaCoSo.objects.filter(nguoi_de_xuat=request.user.username, is_read=False).exclude(trang_thai='ChoDuyet')
        count = unread_proposals.count()
        
        # Determine the latest status for the message if any
        latest_status = None
        if count > 0:
            latest = unread_proposals.order_by('-ngay_cap_nhat').first()
            if latest:
                latest_status = latest.trang_thai
        
        return JsonResponse({'success': True, 'count': count, 'role': 'partner', 'latest_status': latest_status})


@login_required
def api_download_base_price_template(request):
    try:
        import openpyxl
        import urllib.request
        import json
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.core.cache import cache

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template"

        # Sheet "Data" for dropdowns
        ws_data = wb.create_sheet(title="Data")
        ws_data.sheet_state = 'hidden'

        provinces_data = cache.get('provinces_data_full')
        if not provinces_data:
            try:
                # Try fetching from API
                req = urllib.request.Request('https://provinces.open-api.vn/api/?depth=2', headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=5) as response:
                    provinces_data = json.loads(response.read().decode())
                cache.set('provinces_data_full', provinces_data, 86400)
            except Exception as e:
                print("Error fetching provinces:", e)
                provinces_data = []

        import re
        import unicodedata

        def clean_name(s):
            # Same logic as what Excel INDIRECT(SUBSTITUTE(...," ","")) will do for districts
            # Actually, Excel INDIRECT requires the exact name.
            # If the user selects "Hà Nội", SUBSTITUTE(..., " ", "") makes it "HàNội"
            # So the named range MUST be exactly "HàNội".
            # We don't remove diacritics here! We just remove spaces.
            # Let's remove all spaces and hyphens.
            return s.replace(" ", "").replace("-", "").replace(".", "")

        province_names = []
        col_idx = 2
        for p in provinces_data:
            p_name = p['name'].replace("Tỉnh ", "").replace("Thành phố ", "")
            province_names.append(p_name)
            
            # Write province districts to Data sheet
            d_names = [d['name'] for d in p['districts']]
            ws_data.cell(row=1, column=col_idx, value=p_name)
            for r_idx, d_name in enumerate(d_names, start=2):
                ws_data.cell(row=r_idx, column=col_idx, value=d_name)
            
            # Create named range for this province's districts
            c_letter = get_column_letter(col_idx)
            ref = f"Data!${c_letter}$2:${c_letter}${len(d_names)+1}"
            c_name = clean_name(p_name)
            # Create named range (openpyxl allows unicode names for named ranges!)
            try:
                wb.create_named_range(c_name, None, ref)
            except Exception as e:
                print("Could not create named range for", c_name, e)
            
            col_idx += 1

        # Write provinces to Data sheet A2:A...
        ws_data.cell(row=1, column=1, value="Provinces")
        for r_idx, p_name in enumerate(province_names, start=2):
            ws_data.cell(row=r_idx, column=1, value=p_name)
        
        # Create named range for all provinces
        wb.create_named_range("AllProvinces", None, f"Data!$A$2:$A${len(province_names)+1}")

        # Setup Template Sheet headers
        ws.append(["CÔNG TY CỔ PHẦN NHẤT PHONG VẬN"])
        ws.append(["BẢNG GIÁ CƠ SỞ"])
        ws.append([])
        
        ws.merge_cells("A1:M1")
        ws.merge_cells("A2:M2")
        
        title_font = Font(size=16, bold=True, color="FF0000")
        subtitle_font = Font(size=14, bold=True, color="0000FF")
        align_center = Alignment(horizontal="center", vertical="center")
        
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = align_center

        row4 = ["Tỉnh nhận", "Huyện nhận", "Tỉnh giao", "Huyện giao", "Loại xe", "1.25T", "2.5T", "3.5T", "5T", "7T", "8T", "9T", "15T", "LTL"]
        
        from core.models import CauHinhLoaiXe
        default_max = {"1.25T": 9, "2.5T": 13, "3.5T": 18, "5T": 23, "7T": 35, "8T": 54, "9T": 40, "15T": 54}
        row5 = ["", "", "", "", "Số khối"]
        for lx in ["1.25T", "2.5T", "3.5T", "5T", "7T", "8T", "9T", "15T"]:
            ch = CauHinhLoaiXe.objects.filter(loai_xe=lx).first()
            if ch and ch.khoi_den is not None:
                row5.append(f"{ch.khoi_den:g}")
            else:
                row5.append(str(default_max[lx]))
        row5.append("LTL")
        ws.append(row4)
        ws.append(row5)

        ws.merge_cells("A4:A5")
        ws.merge_cells("B4:B5")
        ws.merge_cells("C4:C5")
        ws.merge_cells("D4:D5")
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for r in [4, 5]:
            for c in range(1, 15):
                cell = ws.cell(row=r, column=c)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border

        # Add Data Validations
        if province_names:
            dv_prov = DataValidation(type="list", formula1="=AllProvinces", allow_blank=True)
            ws.add_data_validation(dv_prov)
            dv_prov.add('A6:A1000')
            dv_prov.add('C6:C1000')

            dv_dist_nhan = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(A6, " ", ""), "-", ""), ".", ""))', allow_blank=True)
            ws.add_data_validation(dv_dist_nhan)
            dv_dist_nhan.add('B6:B1000')

            dv_dist_giao = DataValidation(type="list", formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(C6, " ", ""), "-", ""), ".", ""))', allow_blank=True)
            ws.add_data_validation(dv_dist_giao)
            dv_dist_giao.add('D6:D1000')
            
        for col_letter in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions['E'].width = 12
        for col_letter in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col_letter].width = 16
            
        # Format empty rows for prices (columns F to M)
        for row in range(6, 1001):
            for col in range(6, 15):
                c = ws.cell(row=row, column=col)
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal='right', vertical='center')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="template_gia_co_so.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Lỗi tạo file mẫu: {str(e)}", status=500)


@login_required
def api_import_base_prices_excel(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Bạn không có quyền cập nhật.'})
        
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'Chưa chọn file.'})
            
        excel_file = request.FILES['file']
        if not excel_file.name.endswith('.xlsx'):
            return JsonResponse({'success': False, 'error': 'Chỉ hỗ trợ file .xlsx'})
            
        import openpyxl
        from django.utils import timezone
        
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active
        
        # Read row 5 for dynamic volume configs
        from core.models import CauHinhLoaiXe
        
        loai_xe_columns = {
            6: "1.25T", 7: "2.5T", 8: "3.5T", 9: "5T",
            10: "7T", 11: "8T", 12: "9T", 13: "15T", 14: "LTL"
        }
        
        loai_xe_map = {}
        default_max = {
            "1.25T": 9, "2.5T": 13, "3.5T": 18, "5T": 23,
            "7T": 35, "8T": 54, "9T": 40, "15T": 54
        }
        prev_max = 0
        
        for col_idx, loai_xe in loai_xe_columns.items():
            if loai_xe == "LTL":
                loai_xe_map[col_idx] = ("LTL", "LTL")
                continue
                
            so_khoi_val = ws.cell(row=5, column=col_idx).value
            cauhinh = CauHinhLoaiXe.objects.filter(loai_xe=loai_xe).first()
            
            if so_khoi_val is not None:
                try:
                    current_max = float(so_khoi_val)
                except ValueError:
                    current_max = cauhinh.khoi_den if cauhinh and cauhinh.khoi_den is not None else default_max.get(loai_xe, prev_max + 5)
            else:
                current_max = cauhinh.khoi_den if cauhinh and cauhinh.khoi_den is not None else default_max.get(loai_xe, prev_max + 5)
                
            if loai_xe == "1.25T":
                khoi_tu = 0
            else:
                khoi_tu = prev_max
                
            CauHinhLoaiXe.objects.update_or_create(
                loai_xe=loai_xe,
                defaults={'khoi_tu': khoi_tu, 'khoi_den': current_max}
            )
            
            def fmt_num(num):
                return f"{num:g}"
                
            if loai_xe == "1.25T":
                so_khoi_str = f"0-{fmt_num(current_max)}"
            else:
                if khoi_tu == current_max:
                    so_khoi_str = f"{fmt_num(current_max)}"
                else:
                    so_khoi_str = f"{fmt_num(khoi_tu)}-{fmt_num(current_max)}"
                    
            loai_xe_map[col_idx] = (loai_xe, so_khoi_str)
            prev_max = current_max
            
        # Update existing records with the new so_khoi strings
        from core.models import GiaCoSo, DeXuatGiaCoSo
        for col_idx, (loai_xe, so_khoi_str) in loai_xe_map.items():
            if loai_xe != "LTL":
                GiaCoSo.objects.filter(loai_xe=loai_xe).update(so_khoi=so_khoi_str)
                DeXuatGiaCoSo.objects.filter(loai_xe=loai_xe).update(so_khoi=so_khoi_str)
        
        updated_count = 0
        created_count = 0
        
        for r in range(6, ws.max_row + 1):
            tinh_nhan = ws.cell(row=r, column=1).value
            huyen_nhan = ws.cell(row=r, column=2).value
            tinh_giao = ws.cell(row=r, column=3).value
            huyen_giao = ws.cell(row=r, column=4).value
            
            if not tinh_nhan or not tinh_giao:
                continue
                
            tinh_nhan = str(tinh_nhan).strip()
            huyen_nhan = clean_huyen(str(huyen_nhan).strip()) if huyen_nhan else ""
            tinh_giao = str(tinh_giao).strip()
            huyen_giao = clean_huyen(str(huyen_giao).strip()) if huyen_giao else ""
            
            # Find or create TuyenXe
            tuyen, created = TuyenXe.objects.get_or_create(
                tinh_nhan=tinh_nhan,
                huyen_nhan=huyen_nhan,
                tinh_giao=tinh_giao,
                huyen_giao=huyen_giao,
                defaults={
                    'ma_tuyen': f"T_{timezone.now().timestamp()}_{r}" # simple unique id
                }
            )
            
            for col_idx, (loai_xe, so_khoi) in loai_xe_map.items():
                price_val = ws.cell(row=r, column=col_idx).value
                if price_val is not None:
                    try:
                        price = float(price_val)
                        if price <= 0: continue
                        
                        gcs, gcs_created = GiaCoSo.objects.get_or_create(
                            tuyen=tuyen,
                            loai_xe=loai_xe,
                            defaults={
                                'so_khoi': so_khoi,
                                'gia_co_so': price,
                                'ngay_ap_dung': timezone.now()
                            }
                        )
                        
                        if gcs_created:
                            created_count += 1
                        else:
                            changed = False
                            if gcs.gia_co_so != price:
                                old_price = gcs.gia_co_so
                                gcs.gia_co_so = price
                                gcs.ngay_ap_dung = timezone.now()
                                gcs._ly_do_doi = "Cập nhật hàng loạt qua Excel"
                                changed = True
                            
                            if gcs.so_khoi != so_khoi:
                                gcs.so_khoi = so_khoi
                                changed = True
                                
                            if changed:
                                gcs.save()
                                updated_count += 1
                    except ValueError:
                        return JsonResponse({'success': False, 'error': f'Lỗi ở dòng {r}, cột {col_idx}: Mức giá "{price_val}" không hợp lệ (phải là số).'})
                        
        return JsonResponse({'success': True, 'updated': updated_count, 'created': created_count})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})


from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from core.models import UserProfile

@login_required
def user_management_list(request):
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        messages.error(request, 'Bạn không có quyền truy cập trang Quản lý tài khoản.')
        return redirect('search_prices')
        
    users = User.objects.select_related('profile').all().order_by('-date_joined')
    return render(request, 'core/user_management.html', {'users': users})

@login_required
def api_create_user(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
        
    try:
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        first_name = data.get('first_name', '')
        last_name = data.get('last_name', '')
        email = data.get('email', '')
        role = data.get('role', 'ThueNgoai')
        
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'error': 'Tên đăng nhập đã tồn tại'})
            
        user = User.objects.create(
            username=username,
            password=make_password(password),
            first_name=first_name,
            last_name=last_name,
            email=email,
            is_staff=True if role in ['Admin', 'ThueNgoai'] else False
        )
        UserProfile.objects.create(user=user, role=role, created_by=request.user)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_update_user(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
        
    try:
        data = json.loads(request.body)
        user = User.objects.get(pk=pk)
        
        if not request.user.is_superuser and hasattr(user, 'profile') and user.profile.role == 'Admin' and user != request.user:
            return JsonResponse({'success': False, 'error': 'Bạn không thể sửa tài khoản Admin khác'})

        user.first_name = data.get('first_name', user.first_name)
        user.last_name = data.get('last_name', user.last_name)
        user.email = data.get('email', user.email)
        
        password = data.get('password')
        if password:
            user.password = make_password(password)
            
        role = data.get('role')
        if role:
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = role
            profile.save()
            user.is_staff = True if role in ['Admin', 'ThueNgoai'] else False
            
        user.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_delete_user(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
        
    try:
        user = User.objects.get(pk=pk)
        
        if user == request.user:
            return JsonResponse({'success': False, 'error': 'Bạn không thể tự xoá tài khoản của chính mình'})
            
        if not request.user.is_superuser:
            if hasattr(user, 'profile'):
                if user.profile.role == 'Admin' and user.profile.created_by != request.user:
                    return JsonResponse({'success': False, 'error': 'Bạn không thể xoá tài khoản Admin khác nếu không phải người tạo'})
                    
        user.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ============================================================
# Quản lý Tài Khoản - Trang riêng (page-based, not modal)
# ============================================================

def _get_core_permissions():
    """Lấy danh sách quyền của app core có tên tiếng Việt."""
    from django.contrib.auth.models import Permission
    from django.contrib.contenttypes.models import ContentType
    vn_names = {
        'add_doitac': 'Thêm Đối tác',
        'change_doitac': 'Sửa Đối tác',
        'delete_doitac': 'Xóa Đối tác',
        'view_doitac': 'Xem Đối tác',
        'add_giacoso': 'Thêm Giá cơ sở',
        'change_giacoso': 'Sửa Giá cơ sở',
        'delete_giacoso': 'Xóa Giá cơ sở',
        'view_giacoso': 'Xem Giá cơ sở',
        'add_dexuatgiacoso': 'Thêm Đề xuất giá cơ sở',
        'change_dexuatgiacoso': 'Sửa Đề xuất giá cơ sở',
        'delete_dexuatgiacoso': 'Xóa Đề xuất giá cơ sở',
        'view_dexuatgiacoso': 'Xem Đề xuất giá cơ sở',
        'add_tuyenxe': 'Thêm Tuyến xe',
        'change_tuyenxe': 'Sửa Tuyến xe',
        'delete_tuyenxe': 'Xóa Tuyến xe',
        'view_tuyenxe': 'Xem Tuyến xe',
    }
    core_ct = ContentType.objects.filter(app_label='core')
    # Chỉ lấy các quyền chính (thêm, sửa, xóa, xem) cho các model quan trọng
    allowed_prefixes = ('add_', 'change_', 'delete_', 'view_')
    important_models = ('doitac', 'giacoso', 'dexuatgiacoso')
    perms = []
    for p in Permission.objects.filter(content_type__in=core_ct).order_by('content_type__model', 'codename'):
        model_name = '_'.join(p.codename.split('_')[1:])
        if model_name not in important_models:
            continue
        p.vn_name = vn_names.get(p.codename, p.name)
        p.action = p.codename.split('_')[0]
        p.model = model_name
        perms.append(p)
    return perms


@login_required
def account_list(request):
    """Trang danh sách tài khoản - chỉ Admin Tổng."""
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ Admin Tổng mới có quyền truy cập Quản lý tài khoản.')
        return redirect('search_prices')
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'core/account_list.html', {
        'users': users,
        'active_count': users.filter(is_active=True).count(),
        'inactive_count': users.filter(is_active=False).count(),
        'admin_count': users.filter(is_superuser=True).count(),
    })


@login_required
def account_create(request):
    """Trang tạo tài khoản mới."""
    from django.contrib.auth.models import Permission
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ Admin Tổng mới có quyền tạo tài khoản.')
        return redirect('search_prices')

    core_permissions = _get_core_permissions()

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        perm_ids = [int(p) for p in request.POST.getlist('permissions[]') if p.isdigit()]

        error = None
        if not username:
            error = 'Tên đăng nhập là bắt buộc.'
        elif not password:
            error = 'Mật khẩu là bắt buộc.'
        elif len(password) < 6:
            error = 'Mật khẩu phải có ít nhất 6 ký tự.'
        elif User.objects.filter(username=username).exists():
            error = f'Tên đăng nhập "{username}" đã tồn tại.'

        if error:
            messages.error(request, error)
            return render(request, 'core/account_form.html', {
                'title': 'Thêm Tài Khoản Mới',
                'core_permissions': core_permissions,
                'user_perms': perm_ids,
            })

        user = User.objects.create(
            username=username,
            password=make_password(password),
            is_active=True,
            is_staff=True,
        )
        if perm_ids:
            perms = Permission.objects.filter(id__in=perm_ids)
            user.user_permissions.set(perms)

        messages.success(request, f'✅ Đã tạo tài khoản "{username}" thành công!')
        return redirect('account_list')

    return render(request, 'core/account_form.html', {
        'title': 'Thêm Tài Khoản Mới',
        'core_permissions': core_permissions,
        'user_perms': [],
    })


@login_required
def account_update(request, pk):
    """Trang chỉnh sửa tài khoản."""
    from django.contrib.auth.models import Permission
    if not request.user.is_superuser:
        messages.error(request, 'Chỉ Admin Tổng mới có quyền sửa tài khoản.')
        return redirect('search_prices')

    edit_user = get_object_or_404(User, pk=pk)
    core_permissions = _get_core_permissions()
    user_perms = list(edit_user.user_permissions.values_list('id', flat=True))

    if request.method == 'POST':
        password = request.POST.get('password', '').strip()
        is_active = request.POST.get('is_active') == '1'
        perm_ids = [int(p) for p in request.POST.getlist('permissions[]') if p.isdigit()]

        if password:
            if len(password) < 6:
                messages.error(request, 'Mật khẩu mới phải có ít nhất 6 ký tự.')
                return render(request, 'core/account_form.html', {
                    'title': 'Cập Nhật Tài Khoản',
                    'edit_user': edit_user,
                    'core_permissions': core_permissions,
                    'user_perms': user_perms,
                })
            edit_user.set_password(password)

        if not edit_user.is_superuser:
            edit_user.is_active = is_active
            perms = Permission.objects.filter(id__in=perm_ids)
            edit_user.user_permissions.set(perms)

        edit_user.save()
        messages.success(request, f'✅ Đã cập nhật tài khoản "{edit_user.username}" thành công!')
        return redirect('account_list')

    return render(request, 'core/account_form.html', {
        'title': 'Cập Nhật Tài Khoản',
        'edit_user': edit_user,
        'core_permissions': core_permissions,
        'user_perms': user_perms,
    })


@login_required
def api_delete_account(request, pk):
    """API xóa tài khoản (dùng cho nút Xóa trên trang account_list)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Không có quyền'})
    try:
        user = get_object_or_404(User, pk=pk)
        if user == request.user:
            return JsonResponse({'success': False, 'error': 'Bạn không thể tự xóa tài khoản của mình'})
        if user.is_superuser:
            return JsonResponse({'success': False, 'error': 'Không thể xóa tài khoản Admin Tổng'})
        user.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        messages.error(request, 'Bạn không có quyền truy cập trang Quản lý tài khoản.')
        return redirect('search_prices')
        
    users = User.objects.select_related('profile').all().order_by('-date_joined')
    return render(request, 'core/user_management.html', {'users': users})

@login_required
def api_create_user(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
        
    try:
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        first_name = data.get('first_name', '')
        last_name = data.get('last_name', '')
        email = data.get('email', '')
        role = data.get('role', 'ThueNgoai')
        
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'error': 'Tên đăng nhập đã tồn tại'})
            
        user = User.objects.create(
            username=username,
            password=make_password(password),
            first_name=first_name,
            last_name=last_name,
            email=email,
            is_staff=True if role in ['Admin', 'ThueNgoai'] else False
        )
        UserProfile.objects.create(user=user, role=role, created_by=request.user)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_update_user(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
        
    try:
        data = json.loads(request.body)
        user = User.objects.get(pk=pk)
        
        # Admin restrictions: ThueNgoai can't update Admin
        if not request.user.is_superuser and hasattr(user, 'profile') and user.profile.role == 'Admin' and user != request.user:
            return JsonResponse({'success': False, 'error': 'Bạn không thể sửa tài khoản Admin khác'})

        user.first_name = data.get('first_name', user.first_name)
        user.last_name = data.get('last_name', user.last_name)
        user.email = data.get('email', user.email)
        
        password = data.get('password')
        if password:
            user.password = make_password(password)
            
        role = data.get('role')
        if role:
            profile, _ = UserProfile.objects.get_or_create(user=user)
            # Only superusers or the creator can change roles maybe? Let's allow Admin to change role
            profile.role = role
            profile.save()
            user.is_staff = True if role in ['Admin', 'ThueNgoai'] else False
            
        user.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_delete_user(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
        
    if not (request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.role == 'Admin')):
        return JsonResponse({'success': False, 'error': 'Permission denied'})
        
    try:
        user = User.objects.get(pk=pk)
        
        if user == request.user:
            return JsonResponse({'success': False, 'error': 'Bạn không thể tự xoá tài khoản của chính mình'})
            
        if not request.user.is_superuser:
            if hasattr(user, 'profile'):
                if user.profile.role == 'Admin' and user.profile.created_by != request.user:
                    return JsonResponse({'success': False, 'error': 'Bạn không thể xoá tài khoản Admin khác nếu không phải người tạo'})
                    
        user.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# ============================================================
# Cấu hình Loại Xe & Số Khối
# ============================================================
import json

@login_required
def settings_vehicle_list(request):
    if not request.user.is_superuser:
        from django.contrib import messages
        messages.error(request, 'Chỉ Admin Tổng mới có quyền truy cập Cấu hình hệ thống.')
        return redirect('search_prices')
    
    from core.models import CauHinhLoaiXe
    settings = CauHinhLoaiXe.objects.all()
    return render(request, 'core/settings_vehicle.html', {'settings': settings})

@login_required
def api_save_vehicle_setting(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
        
    try:
        data = json.loads(request.body)
        setting_id = data.get('id')
        loai_xe = data.get('loai_xe', '').strip()
        khoi_tu = data.get('khoi_tu')
        khoi_den = data.get('khoi_den')
        
        if not loai_xe:
            return JsonResponse({'success': False, 'error': 'Loại xe không được để trống'})
            
        try:
            khoi_tu = float(khoi_tu) if khoi_tu is not None and str(khoi_tu).strip() != '' else None
        except ValueError:
            khoi_tu = None
            
        try:
            khoi_den = float(khoi_den) if khoi_den is not None and str(khoi_den).strip() != '' else None
        except ValueError:
            khoi_den = None

        from core.models import CauHinhLoaiXe
        if setting_id:
            obj = CauHinhLoaiXe.objects.get(pk=setting_id)
            obj.loai_xe = loai_xe
            obj.khoi_tu = khoi_tu
            obj.khoi_den = khoi_den
            obj.save()
        else:
            if CauHinhLoaiXe.objects.filter(loai_xe=loai_xe).exists():
                return JsonResponse({'success': False, 'error': 'Loại xe này đã tồn tại'})
            CauHinhLoaiXe.objects.create(
                loai_xe=loai_xe,
                khoi_tu=khoi_tu,
                khoi_den=khoi_den
            )
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_delete_vehicle_setting(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False})
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    try:
        from core.models import CauHinhLoaiXe
        CauHinhLoaiXe.objects.filter(pk=pk).delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def api_get_vehicle_settings(request):
    try:
        from core.models import CauHinhLoaiXe
        settings = CauHinhLoaiXe.objects.all()
        data = {s.loai_xe: s.get_so_khoi() for s in settings}
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
